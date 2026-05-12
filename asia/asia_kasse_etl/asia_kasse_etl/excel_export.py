from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .config import TARGET_COLUMNS, XL_EURO_NUM_FMT
from .models import BuchungRow
from .utils import norm_header


def _load_or_create_workbook(output_xlsx: Path):
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    if output_xlsx.exists():
        return load_workbook(output_xlsx)

    workbook = Workbook()
    if workbook.sheetnames == ["Sheet"]:
        workbook.remove(workbook["Sheet"])
    return workbook


def _write_rows_sheet(workbook, rows: Iterable[BuchungRow], sheet_name: str) -> int:
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])

    worksheet = workbook.create_sheet(sheet_name)
    worksheet.append(TARGET_COLUMNS)

    row_count = 0
    for row in rows:
        worksheet.append(
            [
                float(row.umsatz_euro),
                row.bu_gkto,
                row.beleg_1,
                row.datum,
                row.kost_1,
                row.bank,
                row.buchungstext,
            ]
        )
        row_count += 1

    if worksheet.max_row >= 2:
        sum_row_idx = worksheet.max_row + 1
        worksheet.cell(row=sum_row_idx, column=1).value = f"=SUM(A2:A{worksheet.max_row})"
        worksheet.cell(row=sum_row_idx, column=1).number_format = XL_EURO_NUM_FMT
        worksheet.cell(row=sum_row_idx, column=7).value = "Summe"
        for col_idx in range(1, len(TARGET_COLUMNS) + 1):
            worksheet.cell(row=sum_row_idx, column=col_idx).font = Font(bold=True)

    for row_idx in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row_idx, column=1).number_format = XL_EURO_NUM_FMT

    for col_idx, col_name in enumerate(TARGET_COLUMNS, start=1):
        max_len = len(col_name)
        for values in worksheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
            value = values[0]
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 60)

    return row_count


def _rename_umsatz_bank_header(workbook) -> None:
    if "Umsatz" not in workbook.sheetnames:
        return
    worksheet = workbook["Umsatz"]
    for cell in worksheet[1]:
        if norm_header(cell.value) == "Kasse":
            cell.value = "Bank"


def _save_workbook(workbook, output_xlsx: Path) -> Path:
    try:
        workbook.save(output_xlsx)
        return output_xlsx
    except PermissionError:
        candidates = [output_xlsx.with_name(output_xlsx.stem + "_new" + output_xlsx.suffix)]
        candidates.extend(
            output_xlsx.with_name(output_xlsx.stem + f"_new_{idx}" + output_xlsx.suffix)
            for idx in range(2, 11)
        )

        for alt_path in candidates:
            try:
                workbook.save(alt_path)
                print(f"WARNING: Could not overwrite (file open?). Saved to: {alt_path}")
                return alt_path
            except PermissionError:
                continue

        raise


def build_workbook(
    buchung_rows: list[BuchungRow],
    allopay_rows: list[BuchungRow],
    final_rows: list[BuchungRow],
    output_xlsx: Path,
) -> tuple[Path, int, int, int]:
    workbook = _load_or_create_workbook(output_xlsx)
    buchung_count = _write_rows_sheet(workbook, buchung_rows, "Buchung")
    allopay_count = _write_rows_sheet(workbook, allopay_rows, "Allopay")
    final_count = _write_rows_sheet(workbook, final_rows, "Final")
    _rename_umsatz_bank_header(workbook)
    saved_path = _save_workbook(workbook, output_xlsx)
    return saved_path, buchung_count, allopay_count, final_count

