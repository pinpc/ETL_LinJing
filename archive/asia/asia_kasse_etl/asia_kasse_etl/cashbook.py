from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .config import BU_GKTO_NEGATIVE, BU_GKTO_POSITIVE, STANDARD_BANK, STANDARD_KOST
from .models import BuchungRow
from .utils import as_text, format_beleg, norm_header, parse_amount, parse_date


def _find_header_row(ws, min_nonempty: int = 4, scan_rows: int = 30) -> int:
    best_row = 1
    best_score = -1

    for row_idx in range(1, min(ws.max_row, scan_rows) + 1):
        values = [norm_header(cell.value) for cell in ws[row_idx]]
        nonempty = sum(1 for value in values if value != "")
        score = nonempty

        wanted = {"datum", "date", "umsatz", "amount", "betrag", "金额", "交易日期", "识别号", "主题"}
        if any(value in wanted or value.lower() in wanted for value in values):
            score += 5
        if any(value == "金额" for value in values):
            score += 5
        if any(value == "交易日期" for value in values):
            score += 5

        raw_values = [cell.value for cell in ws[row_idx]]
        numericish = sum(
            1
            for value in raw_values
            if isinstance(value, (int, float, Decimal, datetime))
            or (isinstance(value, str) and value.strip().lstrip("-").replace(".", "").replace(",", "").isdigit())
        )
        if numericish >= 2:
            score -= 4

        if score > best_score and nonempty >= min_nonempty:
            best_score = score
            best_row = row_idx

    return best_row


def _map_columns(headers: list[str]) -> dict[str, int]:
    exact = {header: idx for idx, header in enumerate(headers)}
    lower = {header.lower(): idx for idx, header in enumerate(headers)}

    def pick(*candidates: str) -> int | None:
        for candidate in candidates:
            if candidate in exact:
                return exact[candidate]
        for candidate in candidates:
            if candidate.lower() in lower:
                return lower[candidate.lower()]
        for candidate in candidates:
            token = candidate.lower()
            for header, idx in lower.items():
                if token in header:
                    return idx
        return None

    mapping: dict[str, int] = {}
    mapping["Umsatz Euro"] = pick("金额", "Umsatz Euro", "Umsatz", "Amount", "Betrag")  # type: ignore[assignment]
    mapping["Beleg 1"] = pick("识别号", "Beleg 1", "Beleg", "Receipt", "Bon")  # type: ignore[assignment]
    mapping["Datum"] = pick("交易日期", "Datum", "Date")  # type: ignore[assignment]
    mapping["主题"] = pick("主题", "Subject", "Betreff")  # type: ignore[assignment]
    return {key: value for key, value in mapping.items() if value is not None}  # type: ignore[misc]


def read_cashbook_rows(input_xlsx: Path, sheet_name: str = "cashbook") -> list[BuchungRow]:
    workbook = load_workbook(input_xlsx, data_only=True)

    wanted = sheet_name.strip()
    name_map = {name.lower(): name for name in workbook.sheetnames}
    resolved = name_map.get(wanted.lower(), wanted)
    if resolved not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}")

    worksheet = workbook[resolved]
    header_row = _find_header_row(worksheet)
    headers = [norm_header(cell.value) for cell in worksheet[header_row]]
    col_map = _map_columns(headers)

    rows: list[BuchungRow] = []
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        values = [cell.value for cell in worksheet[row_idx]]
        if all(value is None or str(value).strip() == "" for value in values):
            continue

        def get(column_name: str) -> Any:
            idx = col_map.get(column_name)
            return values[idx] if idx is not None and idx < len(values) else None

        amount = parse_amount(get("Umsatz Euro"))
        datum = parse_date(get("Datum"))
        if datum == "" and amount == 0:
            continue

        beleg = format_beleg(get("Beleg 1"))
        subject = as_text(get("主题"))
        buchungstext = "AllO Pay" if beleg.upper().startswith("Z") else subject
        bu_gkto = BU_GKTO_POSITIVE if amount >= 0 else BU_GKTO_NEGATIVE

        rows.append(
            BuchungRow(
                umsatz_euro=amount,
                bu_gkto=bu_gkto,
                beleg_1=beleg,
                datum=datum,
                kost_1=STANDARD_KOST,
                bank=STANDARD_BANK,
                buchungstext=buchungstext,
            )
        )

    return rows

