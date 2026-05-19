"""Bank service placeholders."""

from __future__ import annotations

import csv
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .asia_legacy import run_asia_bank
from ..parser.registry import CsvParser, ParserRegistry
from ..rule_engine.registry import IdentityRule, RulePipeline, RuleSetRegistry
from ..shared.models import ParseRequest, ProcessedTransaction, RuleContext
from ..tenant.models import TenantContext
from ..tenant.service import TenantResolver, resolve_option_path, resolve_option_str
from .interfaces import BankRunRequest, IBankService
from .jupiter_legacy import run_jupiter_bank


class BankService(IBankService):
    """Bank orchestration with tenant-specific legacy adapters."""

    def __init__(self, tenant_resolver: TenantResolver | None = None) -> None:
        self._tenant_resolver = tenant_resolver or TenantResolver()
        parser_registry = ParserRegistry()
        parser_registry.register("csv", CsvParser())
        self._parser_registry = parser_registry

        rule_registry = RuleSetRegistry()
        rule_registry.register("*", "bank", [IdentityRule()])
        self._rule_pipeline = RulePipeline(rule_registry)

    def run(self, request: BankRunRequest) -> list[ProcessedTransaction]:
        tenant_id = request.tenant_id.strip().lower()
        tenant_context = self._tenant_resolver.resolve(tenant_id)
        resolved_request = _resolve_tenant_bank_request(request, tenant_context)

        if tenant_id == "asia":
            run_asia_bank(resolved_request)
            rows = _load_processed_from_bank_workbook(resolved_request.output_path, tenant_id)
            _write_bank_canonical_json(resolved_request.output_path, rows)
            return rows
        if tenant_id == "jupiter":
            run_jupiter_bank(resolved_request)
            rows = _load_processed_from_bank_workbook(resolved_request.output_path, tenant_id)
            _write_bank_canonical_json(resolved_request.output_path, rows)
            return rows

        # Fallback path: generic minimal parser -> rules -> output pipeline.
        parse_request = ParseRequest(
            tenant_id=resolved_request.tenant_id,
            source_type="csv",
            source_path=_resolve_source_file(resolved_request.source_dir),
        )
        parsed_rows = self._parser_registry.parse(parse_request)
        context = RuleContext(tenant_id=resolved_request.tenant_id, module_name="bank")
        processed_rows = self._rule_pipeline.run(parsed_rows, context)
        _write_output_csv(resolved_request.output_path, processed_rows)
        _write_bank_canonical_json(resolved_request.output_path, processed_rows)
        return processed_rows


def _resolve_source_file(source_dir: Path) -> Path:
    if source_dir.is_file():
        return source_dir

    csv_files = sorted(source_dir.glob("*.csv"))
    if not csv_files:
        raise FileNotFoundError(f"No CSV input file found in: {source_dir}")
    return csv_files[0]


def _write_output_csv(output_path: Path, rows) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=[
                "tenant_id",
                "module_name",
                "amount",
                "booking_date",
                "booking_text",
                "bu_gkto",
                "beleg_1",
            ],
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "tenant_id": row.tenant_id,
                    "module_name": row.module_name,
                    "amount": row.amount,
                    "booking_date": row.booking_date,
                    "booking_text": row.booking_text,
                    "bu_gkto": row.bu_gkto,
                    "beleg_1": row.beleg_1,
                }
            )


def _resolve_tenant_bank_request(
    request: BankRunRequest,
    tenant_context: TenantContext,
) -> BankRunRequest:
    return BankRunRequest(
        tenant_id=request.tenant_id,
        source_dir=request.source_dir,
        output_path=request.output_path,
        statement_pdf=request.statement_pdf or resolve_option_path(tenant_context, "bank_statement_pdf"),
        agenda_file=request.agenda_file or resolve_option_path(tenant_context, "bank_agenda_file"),
        sqlite_output_path=request.sqlite_output_path
        or resolve_option_path(tenant_context, "bank_sqlite_output_path"),
        excel_title=request.excel_title or resolve_option_str(tenant_context.options, "bank_excel_title"),
    )


def _load_processed_from_bank_workbook(output_path: Path, tenant_id: str) -> list[ProcessedTransaction]:
    if not output_path.exists():
        raise FileNotFoundError(f"Bank output workbook not found: {output_path}")

    workbook = load_workbook(output_path, data_only=True)
    try:
        sheet_name = _detect_statement_sheet(workbook.sheetnames)
        if sheet_name is None:
            raise ValueError(
                f"Could not detect bank statement sheet in workbook: {output_path}. "
                f"Available sheets: {workbook.sheetnames}"
            )

        worksheet = workbook[sheet_name]
        header_row = _find_header_row(worksheet)
        if header_row is None:
            raise ValueError(
                f"Could not find expected headers in bank sheet '{sheet_name}' of {output_path}."
            )

        rows: list[ProcessedTransaction] = []
        for values in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not values:
                continue
            amount = values[0] if len(values) > 0 else None
            if not isinstance(amount, (int, float)):
                continue
            booking_text = str(values[6]).strip() if len(values) > 6 and values[6] is not None else ""
            if booking_text.upper() in {"TOTAL", "GESAMT"}:
                continue

            rows.append(
                ProcessedTransaction(
                    tenant_id=tenant_id,
                    module_name="bank",
                    amount=float(amount),
                    booking_date=_as_date_text(values[3] if len(values) > 3 else None),
                    booking_text=booking_text,
                    bu_gkto=str(values[1]).strip() if len(values) > 1 and values[1] is not None else "",
                    beleg_1=str(values[2]).strip() if len(values) > 2 and values[2] is not None else "",
                )
            )

        if not rows:
            raise ValueError(
                f"No transaction rows parsed from bank workbook '{output_path}' "
                f"(sheet '{sheet_name}')."
            )
        return rows
    except Exception as exc:
        _write_bank_parse_diagnostics(output_path, workbook, tenant_id, exc)
        raise
    finally:
        workbook.close()


def _detect_statement_sheet(sheet_names: list[str]) -> str | None:
    for candidate in ("Kontoauszug", "Buchungen", "Konto Jupiter"):
        if candidate in sheet_names:
            return candidate
    return sheet_names[0] if sheet_names else None


def _find_header_row(worksheet) -> int | None:
    for row_idx in range(1, min(10, worksheet.max_row) + 1):
        cell_values = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in worksheet[row_idx]]
        if "umsatz euro" in cell_values and "buchungstext" in cell_values:
            return row_idx
    return None


def _as_date_text(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip() if value is not None else ""


def _write_bank_canonical_json(output_path: Path, rows: list[ProcessedTransaction]) -> None:
    json_path = output_path.with_suffix(".processed.json")
    json_path.parent.mkdir(parents=True, exist_ok=True)
    payload = [
        {
            "tenant_id": row.tenant_id,
            "module_name": row.module_name,
            "amount": row.amount,
            "booking_date": row.booking_date,
            "booking_text": row.booking_text,
            "bu_gkto": row.bu_gkto,
            "beleg_1": row.beleg_1,
        }
        for row in rows
    ]
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _write_bank_parse_diagnostics(output_path: Path, workbook, tenant_id: str, exc: Exception) -> None:
    diagnostics_path = output_path.with_suffix(".parse_diagnostics.json")
    diagnostics_path.parent.mkdir(parents=True, exist_ok=True)

    details: dict[str, Any] = {
        "tenant_id": tenant_id,
        "output_path": str(output_path),
        "error_type": type(exc).__name__,
        "error_message": str(exc),
        "sheet_names": list(workbook.sheetnames),
        "sheet_previews": {},
    }

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        preview: list[list[str]] = []
        max_rows = min(5, worksheet.max_row)
        max_cols = min(7, worksheet.max_column)
        for values in worksheet.iter_rows(
            min_row=1,
            max_row=max_rows,
            min_col=1,
            max_col=max_cols,
            values_only=True,
        ):
            preview.append(["" if value is None else str(value).strip() for value in values])
        details["sheet_previews"][sheet_name] = preview

    diagnostics_path.write_text(json.dumps(details, ensure_ascii=False, indent=2), encoding="utf-8")

