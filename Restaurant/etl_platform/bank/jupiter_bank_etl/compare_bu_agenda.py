"""Vergleich BU Gkto: ETL-Ergebnis vs. Jupiter_Agenda_Bank (SOLL).

Wird nur ausgeführt, wenn ein Agenda-Pfad uebergeben wird (z. B. --agenda in launch.json).
"""

from __future__ import annotations

import os
import sys
from collections import defaultdict
from datetime import datetime

from openpyxl import load_workbook


def read_rows(path: str, sheet: str) -> list[tuple]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = []
    for r in range(2, ws.max_row + 1):
        amt = ws.cell(r, 1).value
        bu = ws.cell(r, 2).value
        txt = ws.cell(r, 7).value
        d = ws.cell(r, 4).value
        if txt == "TOTAL" or amt is None or isinstance(amt, str):
            continue
        if isinstance(d, datetime):
            d = d.date()
        bu_s = str(bu).strip() if bu not in (None, "") else ""
        rows.append((round(float(amt), 2), bu_s, str(txt or "").strip(), d))
    wb.close()
    return rows


def _compare_sheet(soll_path: str, ist_path: str, ist_sheet: str, soll_sheet: str = "Final") -> None:
    soll = read_rows(soll_path, soll_sheet)
    ist = read_rows(ist_path, ist_sheet)
    print(f"\n=== IST [{ist_sheet}] {len(ist)} Zeilen | SOLL [{soll_sheet}] {len(soll)} Zeilen ===")

    soll_by_amt: dict[float, list] = defaultdict(list)
    for row in soll:
        soll_by_amt[row[0]].append(row)

    mism = []
    extra_ist = []
    missing_ist = []
    matched = 0
    used: set[tuple] = set()

    for ir in ist:
        amt, ibu, itxt, idt = ir
        cands = soll_by_amt.get(amt, [])
        best = None
        best_score = -1
        for j, sr in enumerate(cands):
            if (amt, j) in used:
                continue
            _, sbu, stxt, sdt = sr
            score = 0
            if sbu == ibu:
                score += 10
            if idt == sdt:
                score += 5
            if itxt[:25] in stxt or stxt[:25] in itxt:
                score += 3
            if score > best_score:
                best_score = score
                best = (j, sr)
        if best and best_score >= 5:
            j, sr = best
            used.add((amt, j))
            matched += 1
            if sr[1] != ibu:
                mism.append((amt, ibu, sr[1], itxt[:55], sr[2][:55]))
        else:
            extra_ist.append((amt, ibu, itxt[:55], idt))

    for amt, cands in soll_by_amt.items():
        for j, sr in enumerate(cands):
            if (amt, j) not in used:
                missing_ist.append((amt, sr[1], sr[2][:55], sr[3]))

    print(
        f"Gematcht: {matched} | BU abweichend: {len(mism)} | "
        f"nur IST: {len(extra_ist)} | nur SOLL: {len(missing_ist)}"
    )

    if mism:
        print("\n--- BU Gkto abweichend (IST vs SOLL) ---")
        for amt, ibu, sbu, itxt, stxt in mism:
            print(f"  {amt:>10.2f}  IST={ibu or '-'}  SOLL={sbu}  | {itxt}")
            if itxt != stxt:
                print(f"             SOLL-Text: {stxt}")

    if extra_ist:
        print("\n--- Nur in IST ---")
        for amt, ibu, itxt, idt in extra_ist[:30]:
            print(f"  {amt:>10.2f}  BU={ibu or '-'}  {idt}  {itxt}")
        if len(extra_ist) > 30:
            print(f"  ... +{len(extra_ist) - 30} weitere")

    if missing_ist:
        print("\n--- Nur in SOLL (fehlt/abweichend in IST) ---")
        for amt, sbu, stxt, sdt in missing_ist[:30]:
            print(f"  {amt:>10.2f}  BU={sbu}  {sdt}  {stxt}")
        if len(missing_ist) > 30:
            print(f"  ... +{len(missing_ist) - 30} weitere")


def run_compare(agenda_path: str, result_path: str) -> None:
    """Vergleicht BU Gkto in Kontoauszug und Final mit der Agenda-Datei."""
    agenda_path = os.path.normpath(agenda_path.strip())
    result_path = os.path.normpath(result_path.strip())

    if not os.path.isfile(agenda_path):
        print(f"WARN Agenda-Vergleich uebersprungen: Datei nicht gefunden: {agenda_path}")
        return
    if not os.path.isfile(result_path):
        print(f"WARN Agenda-Vergleich uebersprungen: Ergebnis nicht gefunden: {result_path}")
        return

    print(f"\n{'=' * 60}")
    print(f"BU-Vergleich Agenda: {os.path.basename(agenda_path)}")
    print(f"Ergebnis:            {os.path.basename(result_path)}")
    print(f"{'=' * 60}")

    for sh in ("Kontoauszug", "Final"):
        _compare_sheet(agenda_path, result_path, sh)


def main(argv: list[str] | None = None) -> None:
    import argparse

    p = argparse.ArgumentParser(description="BU Gkto: ETL-Ergebnis vs. Agenda (SOLL)")
    p.add_argument("--agenda", required=True, help="Pfad zur Agenda-Excel (SOLL)")
    p.add_argument("--out", required=True, help="Pfad zur ETL-Ergebnis-Excel (IST)")
    args = p.parse_args(argv)
    run_compare(args.agenda, args.out)


if __name__ == "__main__":
    main(sys.argv[1:])
