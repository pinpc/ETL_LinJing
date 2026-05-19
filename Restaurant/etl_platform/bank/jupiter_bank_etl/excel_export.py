"""Excel-Workbook (Kontoauszug, Allopay, Final)."""

import re
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .config import XL_EURO_NUM_FMT, XL_FONT, XL_FONT_BOLD
from .expansion import expand_transaction

SHEET_WOLT = "Wolt"
SHEET_ZHOU = "Zhou"


def build_final_sheet(wb: Workbook, bank: str, kost: str, rechnung_map: dict) -> None:
    ws_konto = wb["Kontoauszug"]
    ws_allo = wb["Allopay"]

    allo_pairs = []
    row = 2
    while row <= ws_allo.max_row:
        amt_val = ws_allo.cell(row=row, column=1).value
        fee_val = ws_allo.cell(row=row + 1, column=1).value if row + 1 <= ws_allo.max_row else None
        dat_val = ws_allo.cell(row=row, column=4).value
        txt_amt = ws_allo.cell(row=row, column=7).value or ""
        txt_fee = ws_allo.cell(row=row + 1, column=7).value or "" if row + 1 <= ws_allo.max_row else ""
        if isinstance(amt_val, (int, float)) and amt_val > 0 and isinstance(fee_val, (int, float)) and fee_val < 0:
            ds_m = re.search(r"(\d{2}\.\d{2}\.\d{2,4})", str(txt_amt))
            ds_val = ds_m.group(1) if ds_m else ""
            allo_pairs.append(
                {
                    "amount": round(float(amt_val), 2),
                    "fee": round(float(fee_val), 2),
                    "net": round(float(amt_val) + float(fee_val), 2),
                    "datum": dat_val,
                    "ds": ds_val,
                    "txt_amount": txt_amt,
                    "txt_fee": txt_fee,
                    "used": False,
                }
            )
            row += 2
        else:
            row += 1

    konto_rows = []
    konto_running = 0.0
    for r in ws_konto.iter_rows(min_row=2, max_row=ws_konto.max_row, values_only=True):
        if r[0] is not None and not isinstance(r[0], str):
            konto_rows.append(list(r))
            if str(r[6] or "") != "TOTAL":
                konto_running += float(r[0])

    ws_final = wb.create_sheet("Final")
    for col in range(1, 8):
        src = ws_konto.cell(row=1, column=col)
        dst = ws_final.cell(row=1, column=col)
        dst.value = src.value
        dst.font = src.font.copy() if src.font else None
        dst.fill = src.fill.copy() if src.fill else None
        dst.alignment = src.alignment.copy() if src.alignment else None
        ws_final.column_dimensions[get_column_letter(col)].width = ws_konto.column_dimensions[
            get_column_letter(col)
        ].width

    def thin():
        t = Side(style="thin", color="CCCCCC")
        return Border(left=t, right=t, top=t, bottom=t)

    L = Alignment(horizontal="left", vertical="center")
    ALNS = [L] * 7
    FMTS = [XL_EURO_NUM_FMT, "@", "@", "DD.MM.YYYY", "@", "@", "@"]

    def write_row(ws, row_idx, vals):
        for col, (v, aln, fmt) in enumerate(zip(vals, ALNS, FMTS), 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.border = thin()
            cell.alignment = aln
            if fmt:
                cell.number_format = fmt
            cell.font = XL_FONT

    def write_total_row(ws, row_idx, total):
        total_cell = ws.cell(row=row_idx, column=1, value=round(total, 2))
        total_cell.font = XL_FONT
        total_cell.border = thin()
        total_cell.alignment = L
        total_cell.number_format = XL_EURO_NUM_FMT

        label_cell = ws.cell(row=row_idx, column=7, value="TOTAL")
        label_cell.font = XL_FONT
        label_cell.border = thin()
        label_cell.alignment = L

    final_idx = 2
    final_running = 0.0
    warnings = []

    for kr in konto_rows:
        betrag = kr[0]
        bu = kr[1]
        datum = kr[3]
        txt = kr[6] or ""
        if str(txt) == "TOTAL":
            continue

        if bu == "1360" and "allO pay" in str(txt) and "Gebühr" not in str(txt):
            target_net = round(float(betrag), 2)
            unused = [p for p in allo_pairs if not p["used"]]

            found = None
            for p in unused:
                if abs(p["net"] - target_net) < 0.02:
                    found = [p]
                    break

            if not found:
                running_net = 0.0
                combo = []
                for p in unused:
                    running_net = round(running_net + p["net"], 2)
                    combo.append(p)
                    if abs(running_net - target_net) < 0.02:
                        found = combo
                        break

            if not found:
                best_combo = None
                best_diff = float("inf")
                for size in range(1, min(6, len(unused) + 1)):
                    running_net = 0.0
                    combo = []
                    for p in unused:
                        running_net = round(running_net + p["net"], 2)
                        combo.append(p)
                        diff = abs(running_net - target_net)
                        if diff < best_diff:
                            best_diff = diff
                            best_combo = list(combo)
                        if len(combo) >= size:
                            break
                if best_diff < 1.00 and best_combo:
                    found = best_combo
                    warnings.append(
                        f"allO pay {betrag:.2f}: kleinste Differenz {best_diff:.2f}€ "
                        f"({len(found)} CSV-Zeile(n)) -> Split trotzdem durchgeführt"
                    )

            if found:
                for p in found:
                    p["used"] = True

                total_amount = round(sum(p["amount"] for p in found), 2)
                total_fee = round(sum(p["fee"] for p in found), 2)
                bu_tag = datum

                ds_list = [p["ds"] for p in found if p.get("ds")]
                if len(ds_list) == 1:
                    ds_label = ds_list[0]
                elif len(ds_list) > 1:
                    first = ds_list[0]
                    last = ds_list[-1]
                    first_short = ".".join(first.split(".")[:2])
                    ds_label = f"{first_short} + {last}"
                else:
                    ds_label = ""

                write_row(
                    ws_final,
                    final_idx,
                    [total_amount, "1360", "01", bu_tag, bank, kost, f"all0pay {ds_label}"],
                )
                final_running += total_amount
                final_idx += 1
                write_row(
                    ws_final,
                    final_idx,
                    [total_fee, "4970", "01", bu_tag, bank, kost, f"all0pay Gebühr {ds_label}"],
                )
                final_running += total_fee
                final_idx += 1
            else:
                write_row(ws_final, final_idx, kr[:7])
                final_running += float(kr[0])
                final_idx += 1
                warnings.append(
                    f"allO pay {betrag:.2f} -> kein Match (Differenz >= 1,00EUR), unverändert"
                )
        else:
            key = round(abs(float(betrag)), 2)
            has_invoice_split = (
                key in rechnung_map
                and isinstance(rechnung_map[key][0], str)
                and rechnung_map[key][0].endswith("_SPLIT")
            )
            has_manual_split = abs(round(float(betrag), 2)) == 391.07 and "METRO SAGT DANKE" in str(txt).upper()

            if has_invoice_split or has_manual_split:
                tx = {"betrag": float(betrag), "bu_tag": datum, "beschreibung": txt}
                expanded_rows = expand_transaction(tx, rechnung_map)
                if len(expanded_rows) > 1:
                    for betrag2, bu2, datum2, txt2 in expanded_rows:
                        write_row(ws_final, final_idx, [betrag2, bu2, "01", datum2, bank, kost, txt2])
                        final_running += betrag2
                        final_idx += 1
                    continue

            write_row(ws_final, final_idx, kr[:7])
            final_running += float(kr[0])
            final_idx += 1

    residual = round(konto_running - final_running, 2)
    if residual:
        note_text = (
            f"NOTIZ: Abstimmungsdifferenz zu Konto Jupiter {residual:+.2f} EUR "
            "(z. B. Rundung in Splits)"
        )
        write_row(ws_final, final_idx, [residual, "", "", None, bank, kost, note_text])
        final_running += residual
        final_idx += 1
        warnings.append(note_text)

    write_total_row(ws_final, final_idx, final_running)
    ws_final.freeze_panes = "A2"

    unused_allo = [p for p in allo_pairs if not p["used"]]
    print(f"   Sheet Final: {final_idx - 2} Zeilen")
    if warnings:
        print(f"   WARN: {len(warnings)} allO pay ohne Match:")
        for w in warnings:
            print(f"     {w}")
    if unused_allo:
        print(f"   WARN: {len(unused_allo)} Allopay-Paare nicht verwendet:")
        for p in unused_allo:
            print(f"     {p['datum']} amount={p['amount']} net={p['net']}")


def build_wolt_sheet(wb: Workbook, wolt_audit: list | None) -> None:
    """Blatt Wolt: je Wolt-PDF eine Zeile mit Aufschluesselung und Final-Hinweis."""
    rows_in = list(wolt_audit or [])
    for sn in list(wb.sheetnames):
        if sn.lower() == SHEET_WOLT.lower():
            wb.remove(wb[sn])
    ws = wb.create_sheet(SHEET_WOLT, 1)

    def thin():
        t = Side(style="thin", color="CCCCCC")
        return Border(left=t, right=t, top=t, bottom=t)

    L = Alignment(horizontal="left", vertical="center")
    headers = [
        "Datei",
        "Status",
        "Hinweis / Grund",
        "Nettoauszahlung",
        "Umsatz 7 %",
        "Rabatt",
        "Vertrieb 7 %",
        "Vertrieb 19 %",
        "Gebühr (PDF)",
        "Leistungszeitraum",
        "Summe Probe U+R+V7+V19-G",
        "Diff zu Nettoausz.",
        "Erwartung Blatt Final",
    ]
    widths = [36, 14, 44, 14, 12, 12, 12, 12, 12, 28, 18, 14, 38]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = XL_FONT_BOLD
        c.alignment = L
        c.border = thin()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 15
    ws.freeze_panes = "A2"

    euro_cols = {4, 5, 6, 7, 8, 9, 11, 12}
    if not rows_in:
        rows_in = [
            {
                "datei": "(keine)",
                "status": "-",
                "grund": "Keine Wolt-PDFs verarbeitet (Dateiname muss wolt enthalten).",
            }
        ]

    r = 2
    for item in rows_in:
        vals = [
            item.get("datei"),
            item.get("status"),
            item.get("grund") or "",
            item.get("nettoauszahlung"),
            item.get("umsatz"),
            item.get("rabatt"),
            item.get("provision"),
            item.get("provision19"),
            item.get("gebühr"),
            item.get("zeitraum") or "",
            item.get("probe_summe"),
            item.get("diff_probe"),
            item.get("erw_final") or "",
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=v if v != "" else None)
            cell.font = XL_FONT
            cell.alignment = L
            cell.border = thin()
            if col in euro_cols and isinstance(v, (int, float)):
                cell.number_format = XL_EURO_NUM_FMT
        r += 1

    print(f"   Sheet {SHEET_WOLT}: {len(rows_in)} Zeilen")


def build_zhou_sheet(wb: Workbook, zhou_audit: list | None) -> None:
    """Blatt Zhou: je PDF eine Zeile mit 7 %/19 %-Aufschluesselung und Sammelzeile."""
    rows_in = list(zhou_audit or [])
    for sn in list(wb.sheetnames):
        if sn.lower() == SHEET_ZHOU.lower():
            wb.remove(wb[sn])
    ws = wb.create_sheet(SHEET_ZHOU, 2)

    def thin():
        t = Side(style="thin", color="CCCCCC")
        return Border(left=t, right=t, top=t, bottom=t)

    L = Alignment(horizontal="left", vertical="center")
    headers = [
        "Datei",
        "Status",
        "Typ",
        "Rechnung Nr.",
        "Hinweis / Grund",
        "Brutto 7 %",
        "Brutto 19 %",
        "Endbetrag",
        "Summe 7+19",
        "Diff zu Endbetrag",
        "Erwartung Blatt Final",
    ]
    widths = [36, 14, 6, 14, 36, 12, 12, 12, 12, 14, 36]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = XL_FONT_BOLD
        c.alignment = L
        c.border = thin()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 15
    ws.freeze_panes = "A2"

    euro_cols = {6, 7, 8, 9, 10}
    if not rows_in:
        rows_in = [
            {
                "datei": "(keine)",
                "status": "-",
                "grund": "Keine Zhou-PDFs verarbeitet (Dateiname muss zhou enthalten).",
            }
        ]

    r = 2
    for item in rows_in:
        vals = [
            item.get("datei"),
            item.get("status"),
            item.get("typ") or "",
            item.get("rechnung_nr") or "",
            item.get("grund") or "",
            item.get("netto_7"),
            item.get("netto_19"),
            item.get("endbetrag"),
            item.get("probe_summe"),
            item.get("diff_probe"),
            item.get("erw_final") or "",
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=v if v != "" else None)
            cell.font = XL_FONT
            cell.alignment = L
            cell.border = thin()
            if col in euro_cols and isinstance(v, (int, float)):
                cell.number_format = XL_EURO_NUM_FMT
        r += 1

    print(f"   Sheet {SHEET_ZHOU}: {len(rows_in)} Zeilen")


def build_workbook(
    all_rows: list[tuple],
    output_path: str,
    bank: str,
    kost: str,
    stripe_rows: list,
    rechnung_map: dict,
    wolt_audit: list | None = None,
    zhou_audit: list | None = None,
) -> tuple[int, float]:
    def thin():
        t = Side(style="thin", color="CCCCCC")
        return Border(left=t, right=t, top=t, bottom=t)

    wb = Workbook()
    ws = wb.active
    ws.title = "Kontoauszug"

    L = Alignment(horizontal="left", vertical="center")
    headers = ["Umsatz Euro", "BU Gkto", "Beleg 1", "Datum", "Bank", "Kost 1", "Buchungstext"]
    widths = [14, 10, 8, 13, 8, 8, 60]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = XL_FONT_BOLD
        c.alignment = L
        c.border = thin()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 15
    ws.freeze_panes = "A2"

    ALNS = [L] * 7
    FMTS = [XL_EURO_NUM_FMT, "@", "@", "DD.MM.YYYY", "@", "@", "@"]

    running = 0.0
    for i, (betrag, bu, datum, text) in enumerate(all_rows, 2):
        vals = [betrag, bu, "01", datum, bank, kost, text]
        for col, (v, aln, fmt) in enumerate(zip(vals, ALNS, FMTS), 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.border = thin()
            cell.alignment = aln
            if fmt:
                cell.number_format = fmt
            cell.font = XL_FONT
        running += betrag

    tr = len(all_rows) + 2
    tc = ws.cell(row=tr, column=1, value=round(running, 2))
    tc.font = XL_FONT
    tc.border = thin()
    tc.alignment = L
    tc.number_format = XL_EURO_NUM_FMT
    lc = ws.cell(row=tr, column=7, value="TOTAL")
    lc.font = XL_FONT
    lc.border = thin()
    lc.alignment = L

    build_wolt_sheet(wb, wolt_audit)
    build_zhou_sheet(wb, zhou_audit)

    if stripe_rows:
        ws2 = wb.create_sheet("Allopay")
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            c2 = ws2.cell(row=1, column=col, value=h)
            c2.font = XL_FONT_BOLD
            c2.alignment = L
            c2.border = thin()
            ws2.column_dimensions[get_column_letter(col)].width = w
        ws2.row_dimensions[1].height = 15
        ws2.freeze_panes = "A2"

        row_idx = 2
        for sr in sorted(stripe_rows, key=lambda x: x["date"] or date.today()):
            bu_tag = sr["date"]
            ds2 = sr["ds"]
            pairs = [
                (sr["amount"], "1360", f"allO pay {ds2}"),
                (round(-sr["fee"], 2), "4970", f"allO pay Gebühr {ds2}"),
            ]
            for betrag2, bu2, txt2 in pairs:
                vals2 = [betrag2, bu2, "01", bu_tag, bank, kost, txt2]
                for col, (v, aln, fmt) in enumerate(zip(vals2, ALNS, FMTS), 1):
                    cell = ws2.cell(row=row_idx, column=col, value=v)
                    cell.border = thin()
                    cell.alignment = aln
                    if fmt:
                        cell.number_format = fmt
                    cell.font = XL_FONT
                row_idx += 1
        print(f"   Sheet Allopay: {len(stripe_rows)} Stripe-Dateien -> {row_idx - 2} Zeilen")

        build_final_sheet(wb, bank, kost, rechnung_map)

    wb.save(output_path)
    return len(all_rows), round(running, 2)
