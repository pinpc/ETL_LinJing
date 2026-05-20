"""Bank service placeholders."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .asia_legacy import AsiaLegacyBankRunner
from .errors import BankErrorCode, BankServiceError
from ..parser.registry import CsvParser, ParserRegistry
from ..rule_engine.registry import IdentityRule, RulePipeline, RuleSetRegistry
from ..rule_engine.interfaces import RuleExplainSummary
from ..shared.artifacts import write_run_meta
from ..shared.export_pipeline import (
    ProcessedExportTargets,
    export_processed_rows,
    sidecar_json_path,
    write_sidecar_json,
)
from ..shared.options import first_defined
from ..shared.tenancy import canonical_tenant_id, list_registered_tenant_ids, register_tenant_runner
from ..shared.models import ParseRequest, ProcessedTransaction, RuleContext
from ..tenant.models import TenantContext
from ..tenant.service import TenantResolver, resolve_option_path, resolve_option_str
from .interfaces import BankPipelineResult, BankRunRequest, IBankService, ILegacyBankRunner
from .jupiter_legacy import JupiterLegacyBankRunner

_DEFAULT_LEGACY_BANK_RUNNERS: dict[str, ILegacyBankRunner] = {
    "asia": AsiaLegacyBankRunner(),
    "jupiter": JupiterLegacyBankRunner(),
}


class BankService(IBankService):
    """Bank orchestration with tenant-specific legacy adapters."""

    def __init__(self, tenant_resolver: TenantResolver | None = None) -> None:
        self._tenant_resolver = tenant_resolver or TenantResolver()
        self._legacy_bank_runners = dict(_DEFAULT_LEGACY_BANK_RUNNERS)
        parser_registry = ParserRegistry()
        parser_registry.register("csv", CsvParser())
        self._parser_registry = parser_registry

        rule_registry = RuleSetRegistry()
        rule_registry.register("*", "bank", [IdentityRule()])
        self._rule_pipeline = RulePipeline(rule_registry)

    def run(self, request: BankRunRequest) -> list[ProcessedTransaction]:
        return self.run_with_result(request).rows

    def run_with_result(self, request: BankRunRequest) -> BankPipelineResult:
        tenant_id = canonical_tenant_id(request.tenant_id)
        try:
            tenant_context = self._tenant_resolver.resolve(tenant_id)
            runner_tenant_id = _resolve_runner_tenant_id(tenant_context, tenant_id)
            resolved_request = _resolve_tenant_bank_request(request, tenant_context)

            legacy_runner = self._legacy_bank_runners.get(runner_tenant_id)
            if legacy_runner is not None:
                rows = _run_legacy_bank_pipeline(legacy_runner, resolved_request, tenant_id)
                return _build_pipeline_result(tenant_id, resolved_request.output_path, rows)

            # Fallback path: generic minimal parser -> rules -> output pipeline.
            parse_request = ParseRequest(
                tenant_id=resolved_request.tenant_id,
                source_type="csv",
                source_path=_resolve_source_file(resolved_request.source_dir),
            )
            parsed_rows = self._parser_registry.parse(parse_request)
            context = RuleContext(tenant_id=resolved_request.tenant_id, module_name="bank")
            trace_result = self._rule_pipeline.run_with_trace(parsed_rows, context)
            processed_rows = trace_result.rows
            _export_bank_outputs(resolved_request.output_path, processed_rows)
            trace_summary_path = _write_rule_trace_summary(resolved_request.output_path, trace_result.explain)
            return _build_pipeline_result(
                tenant_id,
                resolved_request.output_path,
                processed_rows,
                rule_trace_summary_path=trace_summary_path,
            )
        except BankServiceError:
            raise
        except FileNotFoundError as exc:
            raise BankServiceError(BankErrorCode.INPUT_MISSING, str(exc)) from exc
        except ValueError as exc:
            raise BankServiceError(BankErrorCode.PARSER_FAILED, str(exc)) from exc
        except RuntimeError as exc:
            message = str(exc)
            if "without producing output workbook" in message:
                raise BankServiceError(BankErrorCode.OUTPUT_NOT_CREATED, message) from exc
            if "Unsupported cashbook tenant" in message or "Unsupported bank tenant" in message:
                raise BankServiceError(BankErrorCode.TENANT_UNSUPPORTED, message) from exc
            raise BankServiceError(BankErrorCode.LEGACY_RUN_FAILED, message) from exc
        except Exception as exc:
            raise BankServiceError(BankErrorCode.UNKNOWN, str(exc)) from exc

    def register_legacy_runner(self, tenant_id: str, runner: ILegacyBankRunner) -> None:
        """Register or override a tenant-specific legacy runner implementation."""
        register_tenant_runner(
            self._legacy_bank_runners,
            tenant_id=tenant_id,
            runner=runner,
            field_name="tenant_id",
        )

    def list_registered_tenants(self) -> list[str]:
        """Return currently registered tenant ids for bank legacy runners."""
        return list_registered_tenant_ids(self._legacy_bank_runners)


def _resolve_source_file(source_dir: Path) -> Path:
    if source_dir.is_file():
        return source_dir

    csv_files = sorted(source_dir.glob("*.csv"))
    if not csv_files:
        raise FileNotFoundError(f"No CSV input file found in: {source_dir}")
    return csv_files[0]


def _run_legacy_bank_pipeline(
    runner: ILegacyBankRunner,
    request: BankRunRequest,
    tenant_id: str,
) -> list[ProcessedTransaction]:
    runner.run(request)
    rows = _load_processed_from_bank_workbook(request.output_path, tenant_id)
    _export_bank_sidecars(request.output_path, rows)
    return rows


def _build_pipeline_result(
    tenant_id: str,
    output_path: Path,
    rows: list[ProcessedTransaction],
    rule_trace_summary_path: Path | None = None,
) -> BankPipelineResult:
    run_meta_path = _write_bank_run_meta(
        tenant_id=tenant_id,
        output_path=output_path,
        row_count=len(rows),
    )
    diagnostics_path = sidecar_json_path(output_path, ".parse_diagnostics.json")
    warnings: list[str] = []
    resolved_diagnostics: Path | None = None
    if diagnostics_path.exists():
        resolved_diagnostics = diagnostics_path
        warnings.append("Parse diagnostics file created; review diagnostics artifact.")
    return BankPipelineResult(
        tenant_id=tenant_id,
        module_name="bank",
        rows=rows,
        output_path=output_path,
        canonical_json_path=output_path.with_suffix(".processed.json"),
        run_meta_path=run_meta_path,
        diagnostics_path=resolved_diagnostics,
        rule_trace_summary_path=rule_trace_summary_path,
        warnings=warnings,
    )


def _resolve_tenant_bank_request(
    request: BankRunRequest,
    tenant_context: TenantContext,
) -> BankRunRequest:
    return BankRunRequest(
        tenant_id=request.tenant_id,
        source_dir=request.source_dir,
        output_path=request.output_path,
        statement_pdf=first_defined(
            request.statement_pdf,
            resolve_option_path(tenant_context, "bank_statement_pdf"),
        ),
        agenda_file=first_defined(
            request.agenda_file,
            resolve_option_path(tenant_context, "bank_agenda_file"),
        ),
        sqlite_output_path=first_defined(
            request.sqlite_output_path,
            resolve_option_path(tenant_context, "bank_sqlite_output_path"),
        ),
        excel_title=first_defined(
            request.excel_title,
            resolve_option_str(tenant_context.options, "bank_excel_title"),
        ),
    )


def _resolve_runner_tenant_id(tenant_context: TenantContext, fallback_tenant_id: str) -> str:
    configured = resolve_option_str(tenant_context.options, "bank_runner_tenant_id")
    if configured:
        return canonical_tenant_id(configured)
    return fallback_tenant_id


def _load_processed_from_bank_workbook(output_path: Path, tenant_id: str) -> list[ProcessedTransaction]:
    if not output_path.exists():
        raise FileNotFoundError(f"Bank output workbook not found: {output_path}")

    workbook = load_workbook(output_path, data_only=True)
    try:
        sheet_name = _detect_statement_sheet(workbook.sheetnames)
        if sheet_name is None:
            raise ValueError(
                f"Could not detect bank statement sheet in workbook: {output_path}. "
                f"Available sheets: {workbook.sheetnames}"
            )

        worksheet = workbook[sheet_name]
        header_row = _find_header_row(worksheet)
        if header_row is None:
            raise ValueError(
                f"Could not find expected headers in bank sheet '{sheet_name}' of {output_path}."
            )

        column_indexes = _extract_column_indexes(worksheet[header_row])
        rows: list[ProcessedTransaction] = []
        for values in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not values:
                continue
            amount = _pick_value(values, column_indexes["amount"], fallback_index=0)
            if not isinstance(amount, (int, float)):
                continue
            booking_text = _as_clean_text(
                _pick_value(values, column_indexes["booking_text"], fallback_index=6)
            )
            if booking_text.upper() in {"TOTAL", "GESAMT"}:
                continue

            rows.append(
                ProcessedTransaction(
                    tenant_id=tenant_id,
                    module_name="bank",
                    amount=float(amount),
                    booking_date=_as_date_text(
                        _pick_value(values, column_indexes["booking_date"], fallback_index=3)
                    ),
                    booking_text=booking_text,
                    bu_gkto=_as_clean_text(
                        _pick_value(values, column_indexes["bu_gkto"], fallback_index=1)
                    ),
                    beleg_1=_as_clean_text(
                        _pick_value(values, column_indexes["beleg_1"], fallback_index=2)
                    ),
                )
            )

        if not rows:
            raise ValueError(
                f"No transaction rows parsed from bank workbook '{output_path}' "
                f"(sheet '{sheet_name}')."
            )
        return rows
    except Exception as exc:
        _write_bank_parse_diagnostics(output_path, workbook, tenant_id, exc)
        raise
    finally:
        workbook.close()


def _detect_statement_sheet(sheet_names: list[str]) -> str | None:
    for candidate in ("Kontoauszug", "Buchungen", "Konto Jupiter"):
        if candidate in sheet_names:
            return candidate
    return sheet_names[0] if sheet_names else None


def _find_header_row(worksheet) -> int | None:
    for row_idx in range(1, min(10, worksheet.max_row) + 1):
        cell_values = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in worksheet[row_idx]]
        if "umsatz euro" in cell_values and "buchungstext" in cell_values:
            return row_idx
    return None


def _as_date_text(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip() if value is not None else ""


def _as_clean_text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _pick_value(values: tuple[Any, ...], index: int | None, fallback_index: int) -> Any:
    effective_index = index if index is not None else fallback_index
    return values[effective_index] if effective_index < len(values) else None


def _extract_column_indexes(header_cells) -> dict[str, int | None]:
    normalized_headers = [
        _normalize_header(cell.value if hasattr(cell, "value") else cell) for cell in header_cells
    ]
    return {
        "amount": _find_header_index(normalized_headers, {"umsatz euro", "umsatz"}),
        "bu_gkto": _find_header_index(normalized_headers, {"bu/gkto", "gkto", "konto"}),
        "beleg_1": _find_header_index(normalized_headers, {"beleg 1", "beleg", "belegnr"}),
        "booking_date": _find_header_index(
            normalized_headers,
            {"buchungstag", "buchungsdatum", "datum"},
        ),
        "booking_text": _find_header_index(
            normalized_headers,
            {"buchungstext", "verwendungszweck", "text"},
        ),
    }


def _find_header_index(headers: list[str], aliases: set[str]) -> int | None:
    alias_set = {_normalize_header(alias) for alias in aliases}
    for idx, header in enumerate(headers):
        if header in alias_set:
            return idx
    return None


def _normalize_header(value: Any) -> str:
    return str(value).strip().lower() if value is not None else ""


def _export_bank_outputs(output_path: Path, rows: list[ProcessedTransaction]) -> None:
    csv_target: Path | None = output_path if output_path.suffix.lower() == ".csv" else None
    export_processed_rows(
        rows,
        ProcessedExportTargets(
            csv_output_path=csv_target,
            json_output_path=output_path.with_suffix(".processed.json"),
        ),
    )


def _export_bank_sidecars(output_path: Path, rows: list[ProcessedTransaction]) -> None:
    export_processed_rows(
        rows,
        ProcessedExportTargets(
            json_output_path=output_path.with_suffix(".processed.json"),
        ),
    )


def _write_bank_parse_diagnostics(output_path: Path, workbook, tenant_id: str, exc: Exception) -> None:
    details: dict[str, Any] = {
        "tenant_id": tenant_id,
        "output_path": str(output_path),
        "error_type": type(exc).__name__,
        "error_message": str(exc),
        "sheet_names": list(workbook.sheetnames),
        "sheet_previews": {},
    }

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        preview: list[list[str]] = []
        max_rows = min(5, worksheet.max_row)
        max_cols = min(7, worksheet.max_column)
        for values in worksheet.iter_rows(
            min_row=1,
            max_row=max_rows,
            min_col=1,
            max_col=max_cols,
            values_only=True,
        ):
            preview.append(["" if value is None else str(value).strip() for value in values])
        details["sheet_previews"][sheet_name] = preview

    write_sidecar_json(output_path, ".parse_diagnostics.json", details)


def _write_bank_run_meta(tenant_id: str, output_path: Path, row_count: int) -> Path:
    return write_run_meta(
        tenant_id=tenant_id,
        module_name="bank",
        output_path=output_path,
        row_count=row_count,
        artifacts={
            "workbook": output_path,
            "processed_json": sidecar_json_path(output_path, ".processed.json"),
            "diagnostics_json": sidecar_json_path(output_path, ".parse_diagnostics.json"),
        },
    )


def _write_rule_trace_summary(
    output_path: Path,
    summary: RuleExplainSummary,
) -> Path:
    return write_sidecar_json(output_path, ".rule_trace_summary.json", summary.to_dict())

