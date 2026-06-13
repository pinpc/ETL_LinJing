"""Final-Blatt: AllOpay-Zeilen aus Buchungen durch Stripe-Splits ersetzen."""

from __future__ import annotations

import itertools
import logging
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .excel_export import UMSATZ_EURO_NUMBERFORMAT, umsatz_zwei_nachkommastellen

logger = logging.getLogger(__name__)

_AMOUNT_TOLERANCE = 0.02

_EDEKA_SPLIT_TEMPLATE: tuple[tuple[str, str], ...] = (
    ("3300", "Edeka WE 7 %"),
    ("3400", "Edeka WE 19 %"),
    ("904250", "Edeka Reinigung"),
)


@dataclass
class _EdekaSheetEntry:
    we7: float
    we19_ohne: float
    reinigung: float
    gesamtbetrag: float
    rechnungsdatum: str = ""
    datei: str = ""
    used: bool = field(default=False, compare=False)


def _cell_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return None


def _load_edeka_entries(wb: Workbook) -> list[_EdekaSheetEntry]:
    """Liest parse-OK Zeilen aus dem EDEKA-Blatt."""
    if "EDEKA" not in wb.sheetnames:
        return []

    ws = wb["EDEKA"]
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        if header:
            headers[str(header).strip()] = col

    required = (
        "WE 7 % Gesamt",
        "WE 19 % ohne Reinigung",
        "Reinigung Wash/Putz",
        "Gesamtbetrag",
        "Parse OK",
    )
    if not all(name in headers for name in required):
        logger.warning("EDEKA-Blatt: erwartete Spalten fehlen – kein Final-Split.")
        return []

    entries: list[_EdekaSheetEntry] = []
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, headers["Parse OK"]).value != "Ja":
            continue

        we7 = _cell_float(ws.cell(row, headers["WE 7 % Gesamt"]).value)
        we19_ohne = _cell_float(ws.cell(row, headers["WE 19 % ohne Reinigung"]).value)
        reinigung = _cell_float(ws.cell(row, headers["Reinigung Wash/Putz"]).value) or 0.0
        gesamt = _cell_float(ws.cell(row, headers["Gesamtbetrag"]).value)
        if we7 is None or we19_ohne is None or gesamt is None:
            continue

        re_datum_col = headers.get("Rechnungsdatum")
        datei_col = headers.get("Datei")
        entries.append(
            _EdekaSheetEntry(
                we7=we7,
                we19_ohne=we19_ohne,
                reinigung=reinigung,
                gesamtbetrag=gesamt,
                rechnungsdatum=str(ws.cell(row, re_datum_col).value or "").strip()
                if re_datum_col
                else "",
                datei=str(ws.cell(row, datei_col).value or "").strip() if datei_col else "",
            )
        )
    return entries


def _find_edeka_match(
    entries: list[_EdekaSheetEntry],
    payment_amount: float,
) -> list[_EdekaSheetEntry] | None:
    """Findet 1 oder 2 unbenutzte EDEKA-Zeilen, deren Gesamtbetrag zur Bankzahlung passt."""
    target = round(abs(payment_amount), 2)

    for entry in entries:
        if entry.used:
            continue
        if abs(entry.gesamtbetrag - target) <= _AMOUNT_TOLERANCE:
            entry.used = True
            return [entry]

    available = [entry for entry in entries if not entry.used]
    for left_idx, left in enumerate(available):
        for right in available[left_idx + 1 :]:
            combined = round(left.gesamtbetrag + right.gesamtbetrag, 2)
            if abs(combined - target) <= _AMOUNT_TOLERANCE:
                left.used = True
                right.used = True
                return [left, right]

    return None


def _unused_edeka_summary(entries: list[_EdekaSheetEntry]) -> str:
    unused = [entry for entry in entries if not entry.used]
    if not unused:
        return "keine unbenutzten EDEKA-Rechnungen mehr verfügbar"
    parts = []
    for entry in unused:
        label = entry.datei or entry.rechnungsdatum or "Rechnung"
        parts.append(f"{label}={entry.gesamtbetrag:.2f} EUR")
    return "verfügbare Gesamtbeträge: " + ", ".join(parts)


def _log_edeka_merge_failure(
    *,
    datum: str,
    betrag: float,
    edeka_entries: list[_EdekaSheetEntry],
    edeka_sheet_present: bool,
) -> None:
    target = round(abs(betrag), 2)
    if not edeka_sheet_present:
        reason = "EDEKA-Blatt fehlt in der Arbeitsmappe"
    elif not edeka_entries:
        reason = "EDEKA-Blatt ohne parse-OK Zeilen"
    else:
        reason = _unused_edeka_summary(edeka_entries)
    logger.warning(
        "Edeka %s (%.2f EUR): Merge mit EDEKA-Blatt fehlgeschlagen "
        "(kein Gesamtbetrag %.2f EUR, einzeln oder als 2er-Summe) – %s; "
        "struktureller Split (nur WE 7%% mit Betrag).",
        datum,
        betrag,
        target,
        reason,
    )


def _invoice_to_bank_amount(raw: float, payment_amount: float) -> float:
    """Rechnungsbetrag (positiv=Aufwand) → Bankzeichen (Lastschrift negativ). Behält Gutschriften."""
    if payment_amount < 0:
        return round(-raw, 2)
    return round(raw, 2)


def _reconcile_edeka_split_parts(
    parts: list[tuple[str, str, float | None]],
    payment_amount: float,
) -> list[tuple[str, str, float | None]]:
    """Stellt sicher, dass die Split-Zeilen exakt der Bankzahlung entsprechen."""
    target = round(payment_amount, 2)
    mutable = [list(part) for part in parts]
    current_total = round(
        sum(part[2] for part in mutable if part[2] is not None),
        2,
    )
    if abs(current_total - target) <= _AMOUNT_TOLERANCE:
        return parts

    diff = round(target - current_total, 2)
    if len(mutable) > 1 and mutable[1][2] is not None:
        mutable[1][2] = round(float(mutable[1][2]) + diff, 2)
        logger.warning(
            "Edeka-Split Summenabgleich: %.2f -> %.2f EUR (WE 19 %% um %+.2f EUR korrigiert)",
            current_total,
            target,
            diff,
        )
    return [tuple(part) for part in mutable]


def _build_edeka_split_parts(
    matches: list[_EdekaSheetEntry],
    payment_amount: float,
) -> list[tuple[str, str, float | None]]:
    we7 = round(sum(entry.we7 for entry in matches), 2)
    we19 = round(sum(entry.we19_ohne for entry in matches), 2)
    reinigung = round(sum(entry.reinigung for entry in matches), 2)

    parts: list[tuple[str, str, float | None]] = [
        (
            _EDEKA_SPLIT_TEMPLATE[0][0],
            _EDEKA_SPLIT_TEMPLATE[0][1],
            _invoice_to_bank_amount(we7, payment_amount),
        ),
        (
            _EDEKA_SPLIT_TEMPLATE[1][0],
            _EDEKA_SPLIT_TEMPLATE[1][1],
            _invoice_to_bank_amount(we19, payment_amount),
        ),
    ]
    if reinigung > _AMOUNT_TOLERANCE:
        parts.append(
            (
                _EDEKA_SPLIT_TEMPLATE[2][0],
                _EDEKA_SPLIT_TEMPLATE[2][1],
                _invoice_to_bank_amount(reinigung, payment_amount),
            )
        )
    else:
        parts.append((_EDEKA_SPLIT_TEMPLATE[2][0], _EDEKA_SPLIT_TEMPLATE[2][1], None))

    return _reconcile_edeka_split_parts(parts, payment_amount)


def _write_split_rows(
    ws_final: Worksheet,
    ws_source: Worksheet,
    src_row: int,
    target_row: int,
    split_parts: list[tuple[str, str, float | None]],
) -> int:
    for bu_gkto, buch_text, amount in split_parts:
        for col in range(1, ws_source.max_column + 1):
            val = ws_source.cell(src_row, col).value
            if col == 1:
                val = umsatz_zwei_nachkommastellen(amount) if amount is not None else None
            elif col == 2:
                val = int(bu_gkto)
            elif col == 7:
                val = buch_text
            ws_final.cell(target_row, col, val)
        target_row += 1
    return target_row


def _is_empty_umsatz(value: Any) -> bool:
    if value is None or value == "":
        return True
    if isinstance(value, (int, float)):
        return abs(float(value)) <= _AMOUNT_TOLERANCE
    return False


def _remove_empty_edeka_final_rows(
    ws_final: Worksheet,
    *,
    start_row: int,
    end_row: int,
) -> tuple[int, int]:
    """Entfernt Edeka-Zeilen mit Umsatz Euro NULL/0 und verdichtet das Final-Blatt."""
    if end_row < start_row:
        return 0, end_row

    write_row = start_row
    removed = 0
    max_col = ws_final.max_column

    for read_row in range(start_row, end_row + 1):
        text = str(ws_final.cell(read_row, 7).value or "").strip()
        amount = ws_final.cell(read_row, 1).value
        if text.startswith("Edeka") and _is_empty_umsatz(amount):
            removed += 1
            continue
        if read_row != write_row:
            for col in range(1, max_col + 1):
                ws_final.cell(write_row, col, ws_final.cell(read_row, col).value)
        write_row += 1

    for clear_row in range(write_row, end_row + 1):
        for col in range(1, max_col + 1):
            ws_final.cell(clear_row, col).value = None

    return removed, write_row - 1


def erstelle_final_blatt(workbook_path: str) -> None:
    """Erstellt Final-Blatt mit intelligenter Ersetzung von AllOpay-Buchungen durch Splits."""
    try:
        wb = load_workbook(workbook_path)
        ws_buchungen = wb["Kontoauszug"]
        ws_allopay = wb["Allopay"]

        if "Final" in wb.sheetnames:
            wb.remove(wb["Final"])
        ws_final = wb.create_sheet("Final")

        for row in [1, 2]:
            for col in range(1, ws_buchungen.max_column + 1):
                ws_final.cell(row, col, ws_buchungen.cell(row, col).value)

        allopay_items: list[dict[str, Any]] = []
        for row in range(2, ws_allopay.max_row + 1):
            datum_cell = ws_allopay.cell(row, 4)
            if not datum_cell.value:
                continue
            if isinstance(datum_cell.value, datetime):
                ap_datum = datum_cell.value.strftime("%d.%m.%Y")
            else:
                ap_datum = str(datum_cell.value).strip()

            betrag_cell = ws_allopay.cell(row, 1)
            betrag = betrag_cell.value
            if not isinstance(betrag, (int, float)):
                continue

            row_values = [
                ws_allopay.cell(row, col).value
                for col in range(1, ws_allopay.max_column + 1)
            ]

            allopay_items.append(
                {
                    "datum": ap_datum,
                    "betrag": betrag,
                    "row_values": row_values,
                    "used": False,
                }
            )

        edeka_entries = _load_edeka_entries(wb)
        edeka_sheet_present = "EDEKA" in wb.sheetnames
        if edeka_entries:
            logger.info("EDEKA-Blatt: %s Rechnung(en) für Final-Split geladen", len(edeka_entries))
        elif edeka_sheet_present:
            logger.warning(
                "EDEKA-Blatt vorhanden, aber keine parse-OK Zeilen – Edeka-Split nur strukturell."
            )

        target_row = 3
        replacements = 0
        edeka_splits = 0
        edeka_merge_failed = 0

        for src_row in range(3, ws_buchungen.max_row + 1):
            src_sum_cell = ws_buchungen.cell(src_row, 1)
            if (
                isinstance(src_sum_cell.value, str)
                and src_sum_cell.value.strip().upper().startswith("=SUM")
            ):
                logger.debug(
                    "Überspringe Quell-Summenzeile in Buchungen: Zeile %s", src_row
                )
                continue

            text_cell = ws_buchungen.cell(src_row, 7)
            text = str(text_cell.value).strip() if text_cell.value else ""
            betrag_cell = ws_buchungen.cell(src_row, 1)
            betrag = betrag_cell.value
            datum_cell = ws_buchungen.cell(src_row, 4)
            if isinstance(datum_cell.value, datetime):
                datum = datum_cell.value.strftime("%d.%m.%Y")
            else:
                datum = str(datum_cell.value).strip() if datum_cell.value else ""

            # Edeka WE/EW: Beträge aus EDEKA-Rechnungsblatt (1 oder 2 PDF-Zeilen → 3 Final-Zeilen).
            if text in ("Edeka WE", "Edeka EW") and isinstance(betrag, (int, float)):
                match = _find_edeka_match(edeka_entries, betrag)
                if match is not None:
                    edeka_splits += 1
                    split_parts = _build_edeka_split_parts(match, betrag)
                    logger.info(
                        "Edeka %s (%.2f EUR): Split aus %s EDEKA-Rechnung(en) -> WE 7%% / WE 19%% / Reinigung",
                        datum,
                        betrag,
                        len(match),
                    )
                    target_row = _write_split_rows(
                        ws_final, ws_buchungen, src_row, target_row, split_parts
                    )
                    continue

                split_parts = [
                    (bu, label, betrag if keep else None)
                    for (bu, label), keep in zip(
                        _EDEKA_SPLIT_TEMPLATE,
                        (True, False, False),
                        strict=True,
                    )
                ]
                edeka_merge_failed += 1
                _log_edeka_merge_failure(
                    datum=datum,
                    betrag=betrag,
                    edeka_entries=edeka_entries,
                    edeka_sheet_present=edeka_sheet_present,
                )
                target_row = _write_split_rows(
                    ws_final, ws_buchungen, src_row, target_row, split_parts
                )
                continue

            if text.startswith("AllOpay") and isinstance(betrag, (int, float)):
                try:
                    buch_date = datetime.strptime(datum, "%d.%m.%Y")
                except ValueError:
                    logger.warning("Ungültiges Datum format: %s", datum)
                    buch_date = datetime.max

                candidates = [
                    item
                    for item in allopay_items
                    if not item["used"]
                    and datetime.strptime(item["datum"], "%d.%m.%Y") < buch_date
                ]
                target = round(betrag, 2)
                best_comb = _find_best_combination(candidates, target)

                if best_comb:
                    replacements += 1
                    logger.info(
                        "Ersetze %s (%.2f €) durch %s Allopay-Zeilen",
                        text,
                        betrag,
                        len(best_comb),
                    )
                    for item in best_comb:
                        item["used"] = True
                        for col_idx, val in enumerate(item["row_values"], 1):
                            if col_idx == 1:
                                val = umsatz_zwei_nachkommastellen(val)
                            ws_final.cell(target_row, col_idx, val)
                        target_row += 1
                    continue
                logger.warning(
                    "Keine passende Kombination für %s (%.2f €)", text, betrag
                )

            for col in range(1, ws_buchungen.max_column + 1):
                val = ws_buchungen.cell(src_row, col).value
                if col == 1:
                    val = umsatz_zwei_nachkommastellen(val)
                ws_final.cell(target_row, col, val)
            target_row += 1

        logger.info("%s Stripe-Buchungen ersetzt", replacements)
        if edeka_splits:
            logger.info("%s Edeka-Zahlung(en) aus EDEKA-Blatt gesplittet", edeka_splits)
        if edeka_merge_failed:
            logger.warning(
                "%s Edeka-Zahlung(en) konnten nicht mit dem EDEKA-Blatt gemergt werden",
                edeka_merge_failed,
            )

        removed_edeka_rows, last_data_row = _remove_empty_edeka_final_rows(
            ws_final, start_row=3, end_row=target_row - 1
        )
        if removed_edeka_rows:
            logger.info(
                "%s Edeka-Zeile(n) ohne Betrag (NULL/0) aus Final entfernt",
                removed_edeka_rows,
            )

        summary_row = last_data_row + 1

        ws_final.cell(summary_row, 7, "GESAMT")
        ws_final.cell(summary_row, 1, f"=SUM(A3:A{last_data_row})")

        for r in range(3, summary_row + 1):
            ca = ws_final.cell(r, 1)
            v = ca.value
            if isinstance(v, (int, float)) or (
                isinstance(v, str) and v.strip().upper().startswith("=SUM")
            ):
                ca.number_format = UMSATZ_EURO_NUMBERFORMAT

        wb.save(workbook_path)
        logger.info("Final-Blatt erstellt: %s", workbook_path)
    except Exception as e:
        logger.error("Fehler beim Erstellen des Final-Blatts: %s", e)
        raise


def _find_best_combination(
    candidates: list[dict[str, Any]],
    target: float,
    tolerance: float = 0.05,
) -> tuple[Any, ...] | None:
    """Findet beste Kombination von Kandidaten mit optimierter Suche (bis max 20 Kombinationen)."""
    if not candidates:
        return None

    limited_candidates = candidates[:20]
    limited_candidates.sort(key=lambda x: x["betrag"], reverse=True)

    for k in range(1, min(len(limited_candidates) + 1, 6)):
        for comb in itertools.combinations(limited_candidates, k):
            total = sum(c["betrag"] for c in comb)
            if abs(total - target) <= tolerance:
                return comb

    return None
