"""Excel-Export: Blätter Kontoauszug, Allopay und Parser Ref."""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook

if __package__ in (None, ""):
    _asia_root = Path(__file__).resolve().parent.parent
    if str(_asia_root) not in sys.path:
        sys.path.insert(0, str(_asia_root))
    from asia_bank_etl.buchungstext_mapping import BUCHUNG_PARSER_RULES
else:
    from .buchungstext_mapping import BUCHUNG_PARSER_RULES

# Excel/XLSX nutzt US-Formatcodes: "," = Tausender, "." = Dezimalstellen — genau 2 Nachkommastellen.
# Darstellung z. B. 30,00 / 1.234,56 übernimmt Excel je nach Windows-Ländereinstellung.
UMSATZ_EURO_NUMBERFORMAT = "#,##0.00"


def umsatz_zwei_nachkommastellen(v: Any) -> Any:
    """Rundet Umsatz auf zwei Dezimalstellen (Excel/Vorschau ohne Gleitkomma-Schweifziffern)."""
    if v is None or v == "":
        return v
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return v


def exportiere_excel(
    rows_buchungen: list[dict[str, Any]],
    rows_allopay: list[dict[str, Any]],
    output_path: str,
    config: dict[str, Any],
    *,
    titel_zeile: str = "Kontoauszug Export  ·  Sparkasse Allgäu  ·  01/2026",
    edeka_rows: list[dict[str, Any]] | None = None,
) -> None:
    """Exportiert Buchungen, Allopay, optional EDEKA-Rechnungen und Parser Ref."""
    wb = Workbook()

    cols = [
        "Umsatz Euro",
        "BU Gkto",
        "Beleg 1",
        "Datum",
        "KOST 1",
        "Bank",
        "Buchungstext",
    ]

    ws = wb.active
    ws.title = "Kontoauszug"

    ws["A1"] = titel_zeile

    for ci, col in enumerate(cols, 1):
        ws.cell(2, ci, col)

    for ri, row in enumerate(rows_buchungen):
        er = ri + 3
        umsatz = umsatz_zwei_nachkommastellen(row["Umsatz Euro"])
        bu = row["BU Gkto"]

        c_u = ws.cell(er, 1, umsatz)
        c_u.number_format = UMSATZ_EURO_NUMBERFORMAT
        ws.cell(er, 2, int(bu) if str(bu).isdigit() else (bu if bu else None))
        ws.cell(er, 3, row["Beleg 1"])
        ws.cell(er, 4, row["Datum"])
        ws.cell(er, 5, config["KOST"])
        ws.cell(er, 6, config["BANK_KONTO"])
        ws.cell(er, 7, row["Buchungstext"])

    last = len(rows_buchungen) + 2
    sr = last + 1
    ws.cell(sr, 7, "GESAMT")
    c_sum = ws.cell(sr, 1, f"=SUM(A3:A{last})")
    c_sum.number_format = UMSATZ_EURO_NUMBERFORMAT

    ws2 = wb.create_sheet("Allopay")
    for ci, col in enumerate(cols, 1):
        ws2.cell(1, ci, col)

    for ri, row in enumerate(rows_allopay):
        er = ri + 2
        umsatz = umsatz_zwei_nachkommastellen(row["Umsatz Euro"])
        bu = row["BU Gkto"]

        c_u2 = ws2.cell(er, 1, umsatz)
        c_u2.number_format = UMSATZ_EURO_NUMBERFORMAT
        ws2.cell(er, 2, int(bu) if str(bu).isdigit() else (bu if bu else None))
        ws2.cell(er, 3, row["Beleg 1"])
        ws2.cell(er, 4, row["Datum"])
        ws2.cell(er, 5, config["STRIPE_KOST"])
        ws2.cell(er, 6, config["STRIPE_BANK"])
        ws2.cell(er, 7, row["Buchungstext"])

    last_stripe = len(rows_allopay) + 1
    sr2 = last_stripe + 1
    ws2.cell(sr2, 7, "GESAMT")
    c_sum2 = ws2.cell(sr2, 1, f"=SUM(A2:A{last_stripe})")
    c_sum2.number_format = UMSATZ_EURO_NUMBERFORMAT

    ws_ref = wb.create_sheet("Parser Ref")
    ref_cols = ["Suchtext (Parser)", "BU Gkto", "Kürzel"]
    for ci, name in enumerate(ref_cols, 1):
        ws_ref.cell(1, ci, name)
    for ri, (suchtext, bu_gkto, kuerzel) in enumerate(
        BUCHUNG_PARSER_RULES, start=2
    ):
        ws_ref.cell(ri, 1, suchtext)
        ws_ref.cell(ri, 2, bu_gkto if bu_gkto is not None else "")
        ws_ref.cell(ri, 3, kuerzel)

    if edeka_rows:
        edeka_cols = [
            "Datei",
            "Rechnung-Nr",
            "Rechnungsdatum",
            "WE 7 % Gesamt",
            "WE 19 % Gesamt",
            "WE 19 % ohne Reinigung",
            "Reinigung Wash/Putz",
            "Gesamtbetrag",
            "Summe 7 % + 19 %",
            "Parse OK",
            "Hinweis",
        ]
        ws_edeka = wb.create_sheet("EDEKA")
        for ci, name in enumerate(edeka_cols, 1):
            ws_edeka.cell(1, ci, name)
        euro_cols = {
            "WE 7 % Gesamt",
            "WE 19 % Gesamt",
            "WE 19 % ohne Reinigung",
            "Reinigung Wash/Putz",
            "Gesamtbetrag",
            "Summe 7 % + 19 %",
        }
        for ri, row in enumerate(edeka_rows, start=2):
            for ci, col in enumerate(edeka_cols, 1):
                val = row.get(col)
                if col in euro_cols and val is not None and val != "":
                    val = umsatz_zwei_nachkommastellen(val)
                    cell = ws_edeka.cell(ri, ci, val)
                    cell.number_format = UMSATZ_EURO_NUMBERFORMAT
                else:
                    ws_edeka.cell(ri, ci, val if val not in ("", None) else None)

    wb.save(output_path)
