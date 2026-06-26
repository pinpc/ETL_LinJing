"""Golden-Master-Vergleich: IST-Excel (Final) vs. SOLL-Agenda-Datei."""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import openpyxl

from .runner import run

# Bekannte EXTRA-Zeilen in IST (Display-Split, fehlende Agenda-Buchungen, …)


def _norm_text(v: str) -> str:
    return re.sub(r"\s+", " ", (v or "").strip())


def _norm_bu(v) -> str:
    return _norm_text(str(v or "")).replace(" ", "")


def _known_extra_key(bu: str, beleg2: str, text: str) -> tuple[str, str, str]:
    return (_norm_bu(bu), beleg2, _norm_text(text))


_KNOWN_EXTRA = {
    _known_extra_key("6855", "2", "Bankgebühren Geldmarktkonto"),
    # DOG Holding: SOLL deckt nur Apr 1–28 ab (6 Buchungen fehlen in Agenda)
    _known_extra_key("1703", "46", "Auszahlung Ramtel12 Darlehen"),
    _known_extra_key("1740", "47", "M. Chen 04 2026"),
    _known_extra_key("1740", "47", "F. Massion 04 2026"),
    _known_extra_key("70104", "47", "ZA-Softwareüberlassung SFirm"),
    _known_extra_key("4970", "47", "Ebics 04 2026"),
    _known_extra_key("4970", "47", "Bankgebühr 04 2026"),
}


@dataclass(frozen=True)
class FinalRow:
    umsatz: Decimal | None
    bu_gkto: str
    beleg1: str
    beleg2: str
    datum: date
    konto: str
    buchungstext: str


def _match_key(row: FinalRow, *, ignore_beleg2: bool = False) -> tuple:
    """Match-Schlüssel: BU + Buchungstext (+ Beleg2, sofern relevant)."""
    b2 = "" if ignore_beleg2 else row.beleg2
    return (_norm_bu(row.bu_gkto), b2, _norm_text(row.buchungstext))


def _norm_beleg1(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() in ("none", ""):
        return ""
    if s.isdigit():
        return str(int(s))
    return s


def _read_final(path: Path, sheet: str | None = None) -> list[FinalRow]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet or wb.sheetnames[0]]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    umsatz_col = idx.get("Umsatz", idx.get("Umsatz Euro"))
    rows: list[FinalRow] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r:
            continue
        u = r[umsatz_col]
        if u is None:
            continue
        bt = r[idx["Buchungstext"]]
        if bt and str(bt).startswith("Summe"):
            continue
        beleg2_raw = r[idx["Beleg2"]]
        beleg2 = str(int(beleg2_raw)) if beleg2_raw is not None else ""
        rows.append(FinalRow(
            umsatz=Decimal(str(u)),
            bu_gkto=str(r[idx["BU Gkto"]] or "").strip(),
            beleg1=_norm_beleg1(r[idx["Beleg1"]]),
            beleg2=beleg2,
            datum=r[idx["Datum"]].date() if isinstance(r[idx["Datum"]], datetime) else r[idx["Datum"]],
            konto=str(r[idx["Konto"]] or "").strip(),
            buchungstext=str(bt or "").strip(),
        ))
    wb.close()
    return rows


def compare_final(
    ist_path: Path,
    soll_path: Path,
    soll_sheet: str | None = None,
    *,
    ignore_beleg2: bool = False,
) -> list[str]:
    """Vergleicht Final-Blatt IST vs. SOLL. Gibt Liste der Abweichungen zurück."""
    ist = _read_final(ist_path, "Final")
    soll = _read_final(soll_path, soll_sheet)

    errors: list[str] = []
    key = lambda r: _match_key(r, ignore_beleg2=ignore_beleg2)
    soll_map = {key(r): r for r in soll}
    ist_map = {key(r): r for r in ist}

    for k, s_row in soll_map.items():
        i_row = ist_map.get(k)
        if i_row is None:
            errors.append(f"FEHLT in IST: BU={s_row.bu_gkto} B2={s_row.beleg2} | {s_row.buchungstext}")
            continue
        if s_row.umsatz is not None and i_row.umsatz is not None:
            if abs(s_row.umsatz - i_row.umsatz) > Decimal("0.01"):
                errors.append(
                    f"Umsatz: SOLL={s_row.umsatz} IST={i_row.umsatz} | {s_row.buchungstext}"
                )
        if s_row.datum != i_row.datum:
            errors.append(f"Datum: SOLL={s_row.datum} IST={i_row.datum} | {s_row.buchungstext}")

    for k, i_row in ist_map.items():
        if k in soll_map:
            continue
        if _known_extra_key(i_row.bu_gkto, i_row.beleg2, i_row.buchungstext) in _KNOWN_EXTRA:
            continue
        errors.append(f"EXTRA in IST: BU={i_row.bu_gkto} B2={i_row.beleg2} | {i_row.buchungstext}")

    ist_sum = sum(r.umsatz for r in ist if r.umsatz is not None)
    konto_sum = sum(r.umsatz for r in _read_final(ist_path, "Kontoauszug") if r.umsatz is not None)
    if abs(ist_sum - konto_sum) > Decimal("0.01"):
        errors.append(f"Summe Final vs Kontoauszug: Final={ist_sum:.2f} Konto={konto_sum:.2f}")

    return errors


@dataclass(frozen=True)
class GoldenCase:
    name: str
    tenant: str
    ist_path: Path
    soll_path: Path
    soll_sheet: str | None = None
    ignore_beleg2: bool = False


_GOLDEN_CASES = (
    GoldenCase(
        "CTM",
        "tenants/ctm",
        Path(r"C:\temp_cursor\DOG\2026\CTM\Fibu 04 2026\Kontoauszug_CTM_04_2026.xlsx"),
        Path(r"C:\temp_cursor\DOG\2026\CTM\Fibu 04 2026\Agenda_CTM_Bank_2026 04.xlsx"),
    ),
    GoldenCase(
        "Ramtel12",
        "tenants/ramtel12",
        Path(r"C:\temp_cursor\DOG\2026\Ramtel12\Fibu 04 2026\Kontoauszug_Ramtel12_04_2026.xlsx"),
        Path(r"C:\temp_cursor\DOG\2026\Ramtel12\Fibu 04 2026\Agenda_Bank_Ramtel12_04 2026.xlsx"),
        "tmp6A99",
    ),
    GoldenCase(
        "DOG Holding",
        "tenants/dog_holding",
        Path(r"C:\temp_cursor\DOG\2026\DOG Holding\Fibu 04 2026\Kontoauszug_DOG_Holding_04_2026.xlsx"),
        Path(r"C:\temp_cursor\DOG\2026\DOG Holding\Fibu 04 2026\Agenda_Bank_Holding 04 2026.xlsx"),
        "tmp8235",
        ignore_beleg2=True,
    ),
)


def run_golden_master() -> int:
    """Führt ETL + Golden-Master aus. Exit-Code 0 = OK."""
    failed = 0
    for case in _GOLDEN_CASES:
        print(f"\n=== {case.name} ===")
        try:
            run(case.tenant)
        except PermissionError:
            print(f"WARN: {case.ist_path.name} gesperrt – vergleiche vorhandene Datei")
        errs = compare_final(
            case.ist_path, case.soll_path, case.soll_sheet,
            ignore_beleg2=case.ignore_beleg2,
        )
        if errs:
            failed += 1
            print(f"FAIL ({len(errs)} Abweichungen):")
            for e in errs[:25]:
                print(f"  - {e}")
            if len(errs) > 25:
                print(f"  ... und {len(errs) - 25} weitere")
        else:
            print("OK – Final entspricht SOLL")
    return failed


if __name__ == "__main__":
    raise SystemExit(run_golden_master())
