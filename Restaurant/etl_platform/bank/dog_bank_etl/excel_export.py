"""Schreibt Agenda-kompatible Excel-Blätter 'Kontoauszug' und 'Final'."""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import NamedTuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SHEET_KONTOAUSZUG = "Kontoauszug"
SHEET_FINAL = "Final"

# Agenda-Spaltenreihenfolge (beide Sheets identisch)
_HEADERS = ("Umsatz", "BU Gkto", "Beleg1", "Beleg2", "Datum", "Konto", "Buchungstext", "Skonto Euro")
_COL_WIDTHS_KONTO = (14, 10, 14, 6, 12, 10, 70, 12)
_COL_WIDTHS_FINAL = (14, 10, 14, 6, 12, 10, 40, 12)

_HEADER_FONT = Font(bold=True, color="000000")
_HEADER_FILL = PatternFill(fill_type=None)
_SUM_FONT    = Font(bold=True)

_DATE_FMT   = "DD.MM.YYYY"
_AMOUNT_FMT = "#,##0.00;-#,##0.00"
_TOP        = Alignment(vertical="top")


def _text_cell(ws, row: int, col: int, value: str):
    """Schreibt eine Text-Zelle (Format @, vertical top)."""
    c = ws.cell(row=row, column=col, value=value)
    c.number_format = "@"
    c.alignment = _TOP
    return c


class ExportRow(NamedTuple):
    umsatz: Decimal        # negativ = Ausgabe, positiv = Einnahme
    bu_gkto: str           # Gegenkonto / BU-Konto
    beleg1: str            # Belegnummer 1 — auto-extrahiert (Kontoauszug-Sheet)
    beleg1_final: str      # Belegnummer 1 — Regel-Override (Final-Sheet)
    beleg2: str            # Kontoauszug-Nummer
    datum: date
    konto: str             # Bankkonto des Tenants (z.B. "1200")
    buchungstext: str      # vollständiger Text — Kontoauszug-Sheet
    buchungstext_kurz: str # ZA/ZE + Kürzel — Final-Sheet
    skonto_euro: Decimal   # meist 0,00


# ---------------------------------------------------------------------------
# Interne Hilfsfunktion: einen Sheet befüllen
# ---------------------------------------------------------------------------

def _fill_sheet(ws, rows: list[ExportRow], col_widths: tuple, use_final: bool) -> None:
    """
    use_final=False → Kontoauszug-Sheet: voller Text, auto-Beleg1, wrap_text
    use_final=True  → Final-Sheet: Kürzel ZA/ZE, Regel-Beleg1, einzeilig
    """
    # Header
    for col, (header, width) in enumerate(zip(_HEADERS, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    # Datenzeilen
    for row_idx, row in enumerate(rows, start=2):
        c1 = ws.cell(row=row_idx, column=1, value=float(row.umsatz))
        c1.number_format = _AMOUNT_FMT
        c1.alignment = _TOP

        _text_cell(ws, row_idx, 2, str(row.bu_gkto))
        _text_cell(ws, row_idx, 3, str(row.beleg1_final if use_final else row.beleg1))
        _text_cell(ws, row_idx, 4, str(row.beleg2))

        c5 = ws.cell(row=row_idx, column=5, value=row.datum)
        c5.number_format = _DATE_FMT
        c5.alignment = _TOP

        _text_cell(ws, row_idx, 6, str(row.konto))

        c7 = ws.cell(row=row_idx, column=7, value=row.buchungstext_kurz if use_final else row.buchungstext)
        c7.alignment = Alignment(wrap_text=(not use_final), vertical="top")

        c8 = ws.cell(row=row_idx, column=8, value=float(row.skonto_euro))
        c8.number_format = _AMOUNT_FMT
        c8.alignment = _TOP

    # Summenzeile
    sum_row = len(rows) + 2
    last_data = sum_row - 1
    for col, label in ((1, "Umsatz"), (8, "Skonto Euro")):
        col_letter = get_column_letter(col)
        c = ws.cell(row=sum_row, column=col,
                    value=f"=SUM({col_letter}2:{col_letter}{last_data})")
        c.number_format = _AMOUNT_FMT
        c.font = _SUM_FONT
        c.alignment = _TOP
    ws.cell(row=sum_row, column=7, value=f"Summe ({len(rows)} Buchungen)").font = _SUM_FONT


# ---------------------------------------------------------------------------
# Öffentliche Funktion
# ---------------------------------------------------------------------------

def write_excel(
    konto_rows: list[ExportRow],
    output_path: str | Path,
    final_rows: list[ExportRow] | None = None,
) -> Path:
    """Schreibt Kontoauszug- und Final-Sheet in eine Excel-Datei.

    konto_rows  → Sheet 'Kontoauszug' (1:1 aus PDF-Transaktionen)
    final_rows  → Sheet 'Final' (kann Split-Zeilen enthalten; None = gleich wie konto_rows)
    """
    if final_rows is None:
        final_rows = konto_rows

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # Sheet 1: Kontoauszug (vollständiger Buchungstext, mehrzeilig)
    ws_konto = wb.active
    ws_konto.title = SHEET_KONTOAUSZUG
    _fill_sheet(ws_konto, konto_rows, _COL_WIDTHS_KONTO, use_final=False)

    # Sheet 2: Final (ZA/ZE + Kürzel, Regel-Beleg1, ggf. aufgesplittet)
    ws_final = wb.create_sheet(title=SHEET_FINAL)
    _fill_sheet(ws_final, final_rows, _COL_WIDTHS_FINAL, use_final=True)

    wb.save(str(output_path))
    return output_path
