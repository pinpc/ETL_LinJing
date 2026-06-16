from __future__ import annotations

import json
import re
import sys
from dataclasses import dataclass, replace
from datetime import datetime
from decimal import Decimal, InvalidOperation
import os
from pathlib import Path
from typing import Any, Iterable

import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

CONFIG_JSON_PATH = Path(__file__).resolve().with_name("jupiter_legacy_config.json")


def _load_config() -> dict:
    return json.loads(CONFIG_JSON_PATH.read_text(encoding="utf-8"))


_CONFIG = _load_config()
_DEFAULTS = _CONFIG["defaults"]
_ACCOUNTS = _CONFIG["accounts"]

TARGET_COLUMNS = list(_CONFIG["target_columns"])
COLUMN_WIDTHS = {key: int(value) for key, value in _CONFIG["column_widths"].items()}

STANDARD_KOST = int(_DEFAULTS["standard_kost"])
STANDARD_BANK = int(_DEFAULTS["standard_bank"])

BU_GKTO_ALLOPAY_EXPENSE = int(_ACCOUNTS["allopay_expense"])
BU_GKTO_BAUMARKT_EXPENSE = int(_ACCOUNTS["baumarkt_expense"])
BU_GKTO_FISCH_FOOD = int(_ACCOUNTS["fisch_food"])
BU_GKTO_PERSONALZIMMER = int(_ACCOUNTS["personalzimmer"])
BU_GKTO_INCOME = int(_ACCOUNTS["income"])
BU_GKTO_ALLOPAY_19 = int(_ACCOUNTS["allopay_19"])
BU_GKTO_ALLOPAY_7 = int(_ACCOUNTS["allopay_7"])
BU_GKTO_TIPS_0 = int(_ACCOUNTS.get("tips_0", 4140))

MERGE_TOLERANCE = Decimal(str(_DEFAULTS["merge_tolerance"]))
XL_EURO_NUM_FMT = str(_DEFAULTS["xl_euro_num_fmt"])
XL_DATE_FMT = "DD.MM.YYYY"
BOOKING_TEXT_ALLO_PAY = "AllO Pay"
BOOKING_TEXT_TIPS = "Trinkgeld"
BOOKING_TEXT_TIPS_0 = "Umsatz 0 %"
BOOKING_TEXT_UMSATZ_19 = "Umsatz 19 %"
BOOKING_TEXT_UMSATZ_7 = "Umsatz 7 %"
BOOKING_TEXT_FISHFOOD_EW_7 = "Fishfood EW 7%"
BOOKING_TEXT_AN_BANK = "an Bank"
BOOKING_TEXT_BANK_DEPOSIT = "Bankeinzahlung"


@dataclass(frozen=True)
class CashbookTransaction:
    datum: str
    einnahmen: Decimal
    ausgaben: Decimal
    buchungstext: str


@dataclass(frozen=True)
class BuchungRow:
    umsatz_euro: Decimal
    bu_gkto: int
    beleg_1: str
    datum: str
    kost_1: int
    bank: int
    buchungstext: str


@dataclass(frozen=True)
class AllopayPdfData:
    datum: str
    z_nummer: str
    umsatz_7: Decimal
    umsatz_19: Decimal
    allopay_payment_sum: Decimal


@dataclass(frozen=True)
class PipelineResult:
    saved_path: Path
    umsatz_count: int
    allopay_count: int
    final_count: int
    pdf_base_dir: Path


def as_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def convert_german_number(value: Any) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    text = str(value).strip().replace("EUR", "").replace("€", "").strip()
    if text == "":
        return Decimal("0")

    text = text.replace(" ", "")
    if "." in text and "," in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        parts = text.split(",")
        if len(parts) == 2 and len(parts[1]) == 2:
            text = text.replace(",", ".")
        else:
            text = text.replace(",", ".")
    elif "." in text:
        if text.count(".") > 1:
            text = text.replace(".", "")
        else:
            parts = text.split(".")
            if len(parts) == 2 and len(parts[1]) != 2:
                text = text.replace(".", "")

    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def parse_money_amount(value: Any) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    text = str(value).strip()
    if text in {"", "-"}:
        return Decimal("0")

    text = text.replace("\ufeff", "").replace("\ufffd", "")
    text = text.replace("EUR", "").replace("€", "")
    text = re.sub(r"\s+", "", text)
    if text == "":
        return Decimal("0")

    neg = False
    if text.startswith("(") and text.endswith(")"):
        neg = True
        text = text[1:-1]
    if text.startswith("-"):
        neg = True
        text = text[1:]

    m = re.search(r"([\d.,]+)", text)
    if not m:
        return Decimal("0")
    core = m.group(1)

    if "," in core and "." in core:
        if core.rfind(".") > core.rfind(","):
            core = core.replace(",", "")
        else:
            core = core.replace(".", "").replace(",", ".")
    elif "," in core:
        parts = core.split(",")
        if len(parts) == 2 and len(parts[1]) <= 2:
            core = parts[0].replace(".", "") + "." + parts[1]
        else:
            core = core.replace(".", "").replace(",", ".")
    elif "." in core:
        if core.count(".") > 1:
            core = core.replace(".", "")
        else:
            parts = core.split(".")
            if len(parts) == 2 and len(parts[1]) != 2:
                core = parts[0] + parts[1]

    try:
        d = Decimal(core)
    except InvalidOperation:
        return Decimal("0")
    return -d if neg else d


def parse_date(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")

    text = str(value).strip()
    if text == "":
        return ""

    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%d.%m.%Y")
        except ValueError:
            pass

    return text


def datum_to_excel_date(value: str) -> datetime | str:
    text = parse_date(value)
    if text == "":
        return ""
    try:
        return datetime.strptime(text, "%d.%m.%Y")
    except ValueError:
        return text


def date_sort_key(value: str) -> tuple[int, int, int, str]:
    text = as_text(value)
    if text == "":
        return (9999, 12, 31, text)

    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y"):
        try:
            dt = datetime.strptime(text, fmt)
            return (dt.year, dt.month, dt.day, text)
        except ValueError:
            pass

    return (9999, 12, 31, text)


def sort_rows_by_date(rows: Iterable[BuchungRow]) -> list[BuchungRow]:
    return sorted(rows, key=lambda row: (date_sort_key(row.datum), row.beleg_1, row.buchungstext))


def _date_group_key(datum: str) -> tuple[int, int, int]:
    return date_sort_key(datum)[:3]


def _buchungstext_lower(text: str) -> str:
    return text.strip().lower()


def trinkgeld_sum_by_date(buchung_rows: Iterable[BuchungRow]) -> dict[tuple[int, int, int], Decimal]:
    totals: dict[tuple[int, int, int], Decimal] = {}
    tips_text = BOOKING_TEXT_TIPS.lower()
    for row in buchung_rows:
        if _buchungstext_lower(row.buchungstext) != tips_text:
            continue
        key = _date_group_key(row.datum)
        totals[key] = totals.get(key, Decimal("0")) + row.umsatz_euro
    return totals


def is_allopay_income_row(row: BuchungRow) -> bool:
    return row.umsatz_euro > 0 and BOOKING_TEXT_ALLO_PAY.lower() in _buchungstext_lower(row.buchungstext)


def _is_fish_buchungstext(text: str) -> bool:
    text_lower = _buchungstext_lower(text)
    return "fisch" in text_lower or "fish" in text_lower


def _is_baumarkt_buchungstext(text: str) -> bool:
    return "baumarkt" in _buchungstext_lower(text)


def _is_personalzimmer_buchungstext(text: str) -> bool:
    return "personalzimmer" in _buchungstext_lower(text)


def normalize_final_row(row: BuchungRow) -> BuchungRow:
    normalized = replace(
        row,
        datum=parse_date(row.datum),
        bu_gkto=get_bu_gkto(row.umsatz_euro, row.buchungstext),
    )
    text_lower = _buchungstext_lower(normalized.buchungstext)

    if _is_fish_buchungstext(text_lower):
        return replace(normalized, buchungstext=BOOKING_TEXT_FISHFOOD_EW_7)
    if text_lower == BOOKING_TEXT_BANK_DEPOSIT.lower():
        return replace(normalized, buchungstext=BOOKING_TEXT_AN_BANK)
    return normalized


def _index_allopay_by_date(
    allopay_rows: list[BuchungRow],
) -> tuple[dict[tuple[int, int, int], list[BuchungRow]], dict[tuple[int, int, int], Decimal]]:
    by_date: dict[tuple[int, int, int], list[BuchungRow]] = {}
    sum_by_date: dict[tuple[int, int, int], Decimal] = {}
    for row in allopay_rows:
        key = _date_group_key(row.datum)
        by_date.setdefault(key, []).append(row)
        sum_by_date[key] = sum_by_date.get(key, Decimal("0")) + row.umsatz_euro
    return by_date, sum_by_date


def _build_safe_allopay_beleg_by_date(allopay_rows: list[BuchungRow]) -> dict[str, str]:
    belege_by_date: dict[str, set[str]] = {}
    for row in allopay_rows:
        beleg = row.beleg_1.strip()
        if beleg == "" or beleg == "Z000":
            continue
        belege_by_date.setdefault(row.datum, set()).add(beleg)

    return {
        datum: next(iter(belege)) if len(belege) == 1 else ""
        for datum, belege in belege_by_date.items()
    }


def _finalize_final_rows(rows: Iterable[BuchungRow]) -> list[BuchungRow]:
    return sort_rows_by_date(normalize_final_row(row) for row in rows)


def build_trinkgeld_final_row(
    trinkgeld_net: Decimal,
    allopay_rows_for_date: list[BuchungRow],
    income_row: BuchungRow,
) -> BuchungRow | None:
    if trinkgeld_net >= Decimal("0"):
        return None

    template = allopay_rows_for_date[0] if allopay_rows_for_date else income_row
    return BuchungRow(
        umsatz_euro=(-trinkgeld_net).quantize(Decimal("0.01")),
        bu_gkto=BU_GKTO_TIPS_0,
        beleg_1=template.beleg_1 or income_row.beleg_1,
        datum=template.datum or income_row.datum,
        kost_1=STANDARD_KOST,
        bank=STANDARD_BANK,
        buchungstext=BOOKING_TEXT_TIPS_0,
    )


def _warn_allopay_merge_failure(message: str) -> None:
    print(f"WARNUNG [Blatt Final]: Allopay-Auflösung nicht mergebar {message}", file=sys.stderr)


def _contains_any(text: str, tokens: tuple[str, ...]) -> bool:
    lowered = text.lower()
    return any(token in lowered for token in tokens)


def _find_header_row(ws, scan_rows: int = 30) -> int:
    best_row = 1
    best_score = -1

    for row_idx in range(1, min(ws.max_row, scan_rows) + 1):
        values = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[row_idx]]
        row_text = " ".join(values).lower()
        nonempty = sum(1 for value in values if value)
        score = nonempty

        if _contains_any(row_text, ("datum", "beleg datum", "einnahmen", "ausgaben", "geschäft", "收入", "支出")):
            score += 8

        if score > best_score and nonempty >= 3:
            best_score = score
            best_row = row_idx

    return best_row


def _map_excel_columns(headers: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}

    for idx, header in enumerate(headers):
        value = header.lower()
        if "datum" in value or "日期" in value:
            mapping.setdefault("datum", idx)
        elif "einnahmen" in value or "rechts" in value or "收入" in value:
            mapping.setdefault("einnahmen", idx)
        elif "ausgaben" in value or "links" in value or "支出" in value:
            mapping.setdefault("ausgaben", idx)
        elif "geschäft" in value or "用途" in value:
            mapping.setdefault("text", idx)

    return mapping


def _should_stop_row(values: list[Any]) -> bool:
    row_text = " ".join(str(value) for value in values if value is not None).lower()
    return _contains_any(row_text, ("summe", "unterschrift", "bestand", "收入总额"))


def _norm_header(value: Any) -> str:
    return as_text(value)


def _resolve_excel_sheet(workbook, sheet_name: str | None):
    if sheet_name and str(sheet_name).strip():
        wanted = str(sheet_name).strip()
        name_map = {name.lower(): name for name in workbook.sheetnames}
        resolved = name_map.get(wanted.lower(), wanted)
        if resolved not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}")
        return workbook[resolved]

    name_map = {name.lower(): name for name in workbook.sheetnames}
    for candidate in ("cashbook", "kasse"):
        if candidate in name_map:
            return workbook[name_map[candidate]]
    return next(workbook[name] for name in workbook.sheetnames if not name.startswith("__"))


def _find_allo_export_header_row(ws, min_nonempty: int = 4, scan_rows: int = 30) -> int:
    best_row = 1
    best_score = -1

    for row_idx in range(1, min(ws.max_row, scan_rows) + 1):
        values = [_norm_header(cell.value) for cell in ws[row_idx]]
        nonempty = sum(1 for value in values if value != "")
        score = nonempty

        wanted = {"datum", "date", "umsatz", "amount", "betrag", "金额", "交易日期", "识别号", "主题"}
        if any(value in wanted or value.lower() in wanted for value in values):
            score += 5
        if any(value == "金额" for value in values):
            score += 5
        if any(value == "交易日期" for value in values):
            score += 5

        raw_values = [cell.value for cell in ws[row_idx]]
        numericish = sum(
            1
            for value in raw_values
            if isinstance(value, (int, float, Decimal, datetime))
            or (
                isinstance(value, str)
                and value.strip().lstrip("-").replace(".", "").replace(",", "").isdigit()
            )
        )
        if numericish >= 2:
            score -= 4

        if score > best_score and nonempty >= min_nonempty:
            best_score = score
            best_row = row_idx

    return best_row


def _map_allo_export_columns(headers: list[str]) -> dict[str, int]:
    exact = {header: idx for idx, header in enumerate(headers)}
    lower = {header.lower(): idx for idx, header in enumerate(headers)}

    def pick(*candidates: str) -> int | None:
        for candidate in candidates:
            if candidate in exact:
                return exact[candidate]
        for candidate in candidates:
            if candidate.lower() in lower:
                return lower[candidate.lower()]
        for candidate in candidates:
            token = candidate.lower()
            for header, idx in lower.items():
                if token in header:
                    return idx
        return None

    mapping: dict[str, int] = {}
    amount_idx = pick("金额", "Umsatz Euro", "Umsatz", "Amount", "Betrag")
    beleg_idx = pick("识别号", "Beleg 1", "Beleg", "Receipt", "Bon")
    datum_idx = pick("交易日期", "Datum", "Date")
    subject_idx = pick("主题", "Subject", "Betreff")
    if amount_idx is not None:
        mapping["amount"] = amount_idx
    if beleg_idx is not None:
        mapping["beleg"] = beleg_idx
    if datum_idx is not None:
        mapping["datum"] = datum_idx
    if subject_idx is not None:
        mapping["subject"] = subject_idx
    return mapping


def _format_beleg(value: Any) -> str:
    text = as_text(value)
    if text == "":
        return ""

    if text.upper().startswith("Z") and any(ch.isdigit() for ch in text):
        digits = "".join(ch for ch in text if ch.isdigit())
        if digits:
            return f"Z{int(digits):03d}"

    digits = "".join(ch for ch in text if ch.isdigit())
    if digits == "":
        return text.upper() if text.upper().startswith("Z") else text
    return f"Z{int(digits):03d}"


def _format_buchungstext_date_short(value: str) -> str:
    text = as_text(value)
    if text == "":
        return ""

    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%d.%m.%y")
        except ValueError:
            pass

    return text


def _worksheet_is_allo_export(worksheet) -> bool:
    header_row = _find_allo_export_header_row(worksheet)
    headers = [_norm_header(cell.value) for cell in worksheet[header_row]]
    col_map = _map_allo_export_columns(headers)
    return "amount" in col_map and "datum" in col_map


def _read_allo_export_rows_from_worksheet(worksheet) -> list[BuchungRow]:
    header_row = _find_allo_export_header_row(worksheet)
    headers = [_norm_header(cell.value) for cell in worksheet[header_row]]
    col_map = _map_allo_export_columns(headers)

    rows: list[BuchungRow] = []
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        values = [cell.value for cell in worksheet[row_idx]]
        if all(value is None or str(value).strip() == "" for value in values):
            continue

        def get(column_name: str) -> Any:
            idx = col_map.get(column_name)
            return values[idx] if idx is not None and idx < len(values) else None

        amount = parse_money_amount(get("amount"))
        datum = parse_date(get("datum"))
        if datum == "" and amount == 0:
            continue

        beleg = _format_beleg(get("beleg"))
        subject = as_text(get("subject"))
        if beleg.upper().startswith("Z"):
            date_suffix = _format_buchungstext_date_short(datum)
            buchungstext = (
                f"{BOOKING_TEXT_ALLO_PAY} {date_suffix}"
                if date_suffix
                else BOOKING_TEXT_ALLO_PAY
            )
        else:
            buchungstext = subject

        rows.append(
            BuchungRow(
                umsatz_euro=amount.quantize(Decimal("0.01")),
                bu_gkto=get_bu_gkto(amount, buchungstext or subject),
                beleg_1=beleg,
                datum=datum,
                kost_1=STANDARD_KOST,
                bank=STANDARD_BANK,
                buchungstext=buchungstext,
            )
        )

    return sort_rows_by_date(rows)


def _read_reference_excel_transactions(worksheet) -> list[CashbookTransaction]:
    header_row = _find_header_row(worksheet)
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in worksheet[header_row]]
    col_map = _map_excel_columns(headers)

    transactions: list[CashbookTransaction] = []
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        values = [cell.value for cell in worksheet[row_idx]]
        if all(value is None or str(value).strip() == "" for value in values):
            continue
        if _should_stop_row(values):
            break

        def get(column_name: str) -> Any:
            idx = col_map.get(column_name)
            return values[idx] if idx is not None and idx < len(values) else None

        datum = parse_date(get("datum"))
        einnahmen = convert_german_number(get("einnahmen"))
        ausgaben = convert_german_number(get("ausgaben"))
        text = str(get("text")).strip() if get("text") is not None else ""

        if datum == "":
            continue
        if einnahmen <= 0 and ausgaben <= 0:
            continue

        transactions.append(
            CashbookTransaction(
                datum=datum,
                einnahmen=einnahmen,
                ausgaben=ausgaben,
                buchungstext=text,
            )
        )

    return transactions


def get_bu_gkto(umsatz: Decimal, text: str) -> int:
    text_lower = _buchungstext_lower(text)

    if umsatz >= 0:
        return BU_GKTO_INCOME
    if text_lower == BOOKING_TEXT_TIPS.lower():
        return BU_GKTO_ALLOPAY_EXPENSE
    if _is_fish_buchungstext(text_lower):
        return BU_GKTO_FISCH_FOOD
    if _is_baumarkt_buchungstext(text_lower):
        return BU_GKTO_ALLOPAY_EXPENSE
    if _is_personalzimmer_buchungstext(text_lower):
        return BU_GKTO_PERSONALZIMMER
    if "allo" in text_lower or text_lower == BOOKING_TEXT_BANK_DEPOSIT.lower():
        return BU_GKTO_ALLOPAY_EXPENSE
    return BU_GKTO_BAUMARKT_EXPENSE


def _build_buchung_rows(transactions: list[CashbookTransaction]) -> list[BuchungRow]:
    rows: list[BuchungRow] = []

    for transaction in transactions:
        if transaction.einnahmen > 0:
            rows.append(
                BuchungRow(
                    umsatz_euro=transaction.einnahmen.quantize(Decimal("0.01")),
                    bu_gkto=get_bu_gkto(transaction.einnahmen, transaction.buchungstext),
                    beleg_1="",
                    datum=transaction.datum,
                    kost_1=STANDARD_KOST,
                    bank=STANDARD_BANK,
                    buchungstext=transaction.buchungstext,
                )
            )
        if transaction.ausgaben > 0:
            negative_amount = (-transaction.ausgaben).quantize(Decimal("0.01"))
            rows.append(
                BuchungRow(
                    umsatz_euro=negative_amount,
                    bu_gkto=get_bu_gkto(negative_amount, transaction.buchungstext),
                    beleg_1="",
                    datum=transaction.datum,
                    kost_1=STANDARD_KOST,
                    bank=STANDARD_BANK,
                    buchungstext=transaction.buchungstext,
                )
            )

    sorted_rows = sort_rows_by_date(rows)
    return [
        BuchungRow(
            umsatz_euro=row.umsatz_euro,
            bu_gkto=row.bu_gkto,
            beleg_1=f"Z{idx:03d}",
            datum=row.datum,
            kost_1=row.kost_1,
            bank=row.bank,
            buchungstext=row.buchungstext,
        )
        for idx, row in enumerate(sorted_rows, start=1)
    ]


def read_cashbook_rows(input_path: Path, sheet_name: str | None = None) -> list[BuchungRow]:
    suffix = input_path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise ValueError(f"Unsupported cashbook format: {input_path.suffix}")

    workbook = load_workbook(input_path, data_only=True)
    try:
        worksheet = _resolve_excel_sheet(workbook, sheet_name)
        if _worksheet_is_allo_export(worksheet):
            return _read_allo_export_rows_from_worksheet(worksheet)
        transactions = _read_reference_excel_transactions(worksheet)
    finally:
        workbook.close()
    return _build_buchung_rows(transactions)


def _table_gross_amount(row: list) -> Decimal:
    if not row:
        return Decimal("0")
    if len(row) >= 4 and row[3] is not None and str(row[3]).strip():
        value = parse_money_amount(row[3])
        if value != 0:
            return value
    for cell in reversed(row[1:]):
        if cell is None:
            continue
        value = parse_money_amount(cell)
        if value != 0:
            return value
    return Decimal("0")


def _find_allopay_pdfs(base_path: Path) -> list[Path]:
    pattern = re.compile(r"jupiter \d{2}-\d{2}-\d{4}\.pdf$", re.IGNORECASE)
    matched: list[Path] = []

    for root, _dirs, files in os.walk(base_path):
        for filename in files:
            if pattern.match(filename):
                matched.append(Path(root) / filename)

    return sorted(matched)


def _extract_allopay_from_pdf(pdf_path: Path) -> AllopayPdfData:
    umsatz_7 = convert_german_number(0)
    umsatz_19 = convert_german_number(0)
    allopay_payment_sum = convert_german_number(0)
    datum = ""
    z_nummer = ""

    with pdfplumber.open(str(pdf_path)) as pdf:
        full_text = ""
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                full_text += text + "\n"

        z_match = re.search(r"Z-Nummer\s*\(Berichtnummer\):\s*(\d+)", full_text, re.IGNORECASE)
        if z_match:
            z_nummer = z_match.group(1)

        generiert_match = re.search(r"Generiert um:\s*(\d{2})-(\d{2})-(\d{4})", full_text)
        if generiert_match:
            datum = f"{generiert_match.group(1)}.{generiert_match.group(2)}.{generiert_match.group(3)}"
        else:
            date_match = re.search(r"jupiter (\d{2})-(\d{2})-(\d{4})\.pdf", pdf_path.name, re.IGNORECASE)
            if date_match:
                datum = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"

        for page in pdf.pages:
            for table in page.extract_tables() or []:
                for row in table or []:
                    if not row or len(row) < 4:
                        continue

                    row_text = " ".join(str(cell) if cell else "" for cell in row)
                    first_cell = str(row[0]).strip().lower() if row[0] else ""
                    gross = _table_gross_amount(row)

                    if "Umsatz 7%" in row_text or "Umsatz 7 %" in row_text:
                        if gross > 0:
                            umsatz_7 = gross
                    if "Umsatz 19%" in row_text or "Umsatz 19 %" in row_text:
                        if gross > 0:
                            umsatz_19 = gross
                    if "allo" in first_cell and "pay" in first_cell:
                        if gross > 0:
                            allopay_payment_sum = max(allopay_payment_sum, gross)

        if umsatz_7 == 0 and umsatz_19 == 0:
            match_7 = re.search(
                r"Umsatz\s*7\s*%\D*([\d.,]+)\D+([\d.,]+)\D+([\d.,]+)",
                full_text,
                re.IGNORECASE,
            )
            if match_7:
                umsatz_7 = parse_money_amount(match_7.group(3))

            match_19 = re.search(
                r"Umsatz\s*19\s*%\D*([\d.,]+)\D+([\d.,]+)\D+([\d.,]+)",
                full_text,
                re.IGNORECASE,
            )
            if match_19:
                umsatz_19 = parse_money_amount(match_19.group(3))

        if allopay_payment_sum == 0:
            payment_match = re.search(
                r"allO\s*Pay\D*([\d.,]+)\D+([\d.,]+)\D+([\d.,]+)",
                full_text,
                re.IGNORECASE,
            )
            if payment_match:
                allopay_payment_sum = parse_money_amount(payment_match.group(3))

    return AllopayPdfData(
        datum=datum,
        z_nummer=z_nummer,
        umsatz_7=umsatz_7,
        umsatz_19=umsatz_19,
        allopay_payment_sum=allopay_payment_sum,
    )


def read_allopay_pdf_data(pdf_base_dir: Path) -> list[AllopayPdfData]:
    data_items: list[AllopayPdfData] = []

    for pdf_path in _find_allopay_pdfs(pdf_base_dir):
        data = _extract_allopay_from_pdf(pdf_path)
        if data.datum != "":
            data_items.append(data)

    return sorted(data_items, key=lambda item: item.datum)


def build_allopay_rows(allopay_pdf_data: list[AllopayPdfData]) -> list[BuchungRow]:
    rows: list[BuchungRow] = []

    for data in allopay_pdf_data:
        beleg = f"Z{str(data.z_nummer).zfill(3)}" if data.z_nummer else "Z000"

        if data.umsatz_19 > 0:
            rows.append(
                BuchungRow(
                    umsatz_euro=data.umsatz_19,
                    bu_gkto=BU_GKTO_ALLOPAY_19,
                    beleg_1=beleg,
                    datum=data.datum,
                    kost_1=STANDARD_KOST,
                    bank=STANDARD_BANK,
                    buchungstext="Umsatz 19 %",
                )
            )

        if data.umsatz_7 > 0:
            rows.append(
                BuchungRow(
                    umsatz_euro=data.umsatz_7,
                    bu_gkto=BU_GKTO_ALLOPAY_7,
                    beleg_1=beleg,
                    datum=data.datum,
                    kost_1=STANDARD_KOST,
                    bank=STANDARD_BANK,
                    buchungstext="Umsatz 7 %",
                )
            )

    return sort_rows_by_date(rows)


def _append_sheet(workbook: Workbook, rows: list[BuchungRow], sheet_name: str) -> int:
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.append(TARGET_COLUMNS)
    use_excel_dates = sheet_name == "Final"

    for row in rows:
        datum_value = datum_to_excel_date(row.datum) if use_excel_dates else row.datum
        worksheet.append(
            [
                float(row.umsatz_euro),
                row.bu_gkto,
                row.beleg_1,
                datum_value,
                row.kost_1,
                row.bank,
                row.buchungstext,
            ]
        )

    if rows:
        worksheet.append(
            [
                float(sum((row.umsatz_euro for row in rows), Decimal("0"))),
                "",
                "",
                "",
                "",
                "",
                "Gesamtbetrag",
            ]
        )

    _apply_excel_formatting(worksheet, use_excel_dates=use_excel_dates)
    return len(rows)


def _apply_excel_formatting(worksheet, *, use_excel_dates: bool = False) -> None:
    max_row = worksheet.max_row
    max_col = len(TARGET_COLUMNS)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx in range(2, max_row + 1):
        amount_cell = worksheet.cell(row=row_idx, column=1)
        if amount_cell.value not in (None, ""):
            amount_cell.number_format = XL_EURO_NUM_FMT

        if use_excel_dates:
            date_cell = worksheet.cell(row=row_idx, column=4)
            if isinstance(date_cell.value, datetime):
                date_cell.number_format = XL_DATE_FMT

        if worksheet.cell(row=row_idx, column=7).value == "Gesamtbetrag":
            for col_idx in range(1, max_col + 1):
                worksheet.cell(row=row_idx, column=col_idx).font = Font(bold=True)

    for col_letter, width in COLUMN_WIDTHS.items():
        worksheet.column_dimensions[col_letter].width = width


def _save_workbook(workbook: Workbook, output_xlsx: Path) -> Path:
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)

    try:
        workbook.save(output_xlsx)
        return output_xlsx
    except PermissionError:
        candidates = [output_xlsx.with_name(output_xlsx.stem + "_new" + output_xlsx.suffix)]
        candidates.extend(
            output_xlsx.with_name(output_xlsx.stem + f"_new_{idx}" + output_xlsx.suffix)
            for idx in range(2, 11)
        )

        for alt_path in candidates:
            try:
                workbook.save(alt_path)
                print(f"WARNING: Could not overwrite (file open?). Saved to: {alt_path}")
                return alt_path
            except PermissionError:
                continue

        raise


def build_workbook(
    umsatz_rows: list[BuchungRow],
    allopay_rows: list[BuchungRow],
    final_rows: list[BuchungRow],
    output_xlsx: Path,
) -> tuple[Path, int, int, int]:
    workbook = Workbook()
    if workbook.sheetnames == ["Sheet"]:
        workbook.remove(workbook["Sheet"])

    umsatz_count = _append_sheet(workbook, umsatz_rows, "Umsatz")
    allopay_count = _append_sheet(workbook, allopay_rows, "Allopay")
    final_count = _append_sheet(workbook, final_rows, "Final")

    saved_path = _save_workbook(workbook, output_xlsx)
    return saved_path, umsatz_count, allopay_count, final_count


class JupiterKasseETL:
    def __init__(self) -> None:
        self.umsatz_rows: list[BuchungRow] = []
        self.allopay_rows: list[BuchungRow] = []
        self.final_rows: list[BuchungRow] = []

    def merge_final_rows(
        self,
        umsatz_rows: list[BuchungRow],
        allopay_rows: list[BuchungRow],
        allopay_payment_sum_by_date: dict[str, Decimal] | None = None,
    ) -> list[BuchungRow]:
        if allopay_payment_sum_by_date is None:
            allopay_payment_sum_by_date = {}

        if not umsatz_rows:
            return []

        if not allopay_rows:
            for row in umsatz_rows:
                if is_allopay_income_row(row):
                    _warn_allopay_merge_failure(
                        f"({row.datum}): keine Allopay-PDF-Zeilen geladen; "
                        f"Kasse bleibt {row.umsatz_euro} EUR."
                    )
            return _finalize_final_rows(replace(row, beleg_1="") for row in umsatz_rows)

        allopay_by_date, allopay_sum_by_date = _index_allopay_by_date(allopay_rows)
        trinkgeld_by_date = trinkgeld_sum_by_date(umsatz_rows)
        safe_allopay_beleg_by_date = _build_safe_allopay_beleg_by_date(allopay_rows)

        final_rows: list[BuchungRow] = []
        merged_allopay_income_keys: set[tuple[int, int, int]] = set()

        for row in umsatz_rows:
            if not is_allopay_income_row(row):
                continue
            key = _date_group_key(row.datum)
            pdf_umsatz_sum = allopay_sum_by_date.get(key)
            if pdf_umsatz_sum is None:
                _warn_allopay_merge_failure(
                    f"({row.datum}): keine PDF-Umsatzzeilen (19 %/7 %) für dieses Datum; "
                    f"Kasse bleibt {row.umsatz_euro} EUR."
                )
                continue

            trinkgeld = trinkgeld_by_date.get(key, Decimal("0"))
            adjusted_income = (row.umsatz_euro + trinkgeld).quantize(Decimal("0.01"))
            diff = abs(pdf_umsatz_sum - adjusted_income)
            if diff <= MERGE_TOLERANCE:
                safe_beleg = safe_allopay_beleg_by_date.get(row.datum, "")
                final_rows.extend(
                    replace(allopay_row, beleg_1=safe_beleg)
                    for allopay_row in allopay_by_date[key]
                )
                trinkgeld_row = build_trinkgeld_final_row(trinkgeld, allopay_by_date[key], row)
                if trinkgeld_row is not None:
                    final_rows.append(trinkgeld_row)
                merged_allopay_income_keys.add(key)
                continue

            tip_hint = (
                f", Trinkgeld {trinkgeld} EUR, netto {adjusted_income} EUR"
                if trinkgeld != Decimal("0")
                else ""
            )
            _warn_allopay_merge_failure(
                f"({row.datum}): Kasse {row.umsatz_euro} EUR{tip_hint} vs. Summe PDF "
                f"({BOOKING_TEXT_UMSATZ_19} + {BOOKING_TEXT_UMSATZ_7}) {pdf_umsatz_sum} EUR "
                f"(|Diff|={diff}, Toleranz {MERGE_TOLERANCE})."
            )

        for row in umsatz_rows:
            key = _date_group_key(row.datum)

            if is_allopay_income_row(row) and key in merged_allopay_income_keys:
                continue

            split_rows = self._split_compound_final_row(
                row,
                allopay_payment_sum_by_date,
                safe_allopay_beleg_by_date,
            )
            if split_rows is not None:
                final_rows.extend(split_rows)
                continue

            final_rows.append(self._assign_final_beleg(row, safe_allopay_beleg_by_date))

        return _finalize_final_rows(final_rows)

    def _split_compound_final_row(
        self,
        row: BuchungRow,
        allopay_payment_sum_by_date: dict[str, Decimal],
        safe_allopay_beleg_by_date: dict[str, str],
    ) -> list[BuchungRow] | None:
        text_lower = _buchungstext_lower(row.buchungstext)
        is_compound_row = row.umsatz_euro < 0 and "+" in row.buchungstext and "allo" in text_lower
        if not is_compound_row:
            return None

        allopay_total = allopay_payment_sum_by_date.get(row.datum)
        if allopay_total is None or allopay_total <= 0:
            return None

        expense_total = abs(row.umsatz_euro)
        allopay_expense = min(expense_total, allopay_total).quantize(Decimal("0.01"))
        remainder_expense = (expense_total - allopay_expense).quantize(Decimal("0.01"))
        safe_allopay_beleg = safe_allopay_beleg_by_date.get(row.datum, "")

        split_rows = [
            replace(
                row,
                umsatz_euro=-allopay_expense,
                bu_gkto=BU_GKTO_ALLOPAY_EXPENSE,
                beleg_1=safe_allopay_beleg,
                buchungstext=f"alloPay {row.datum}",
            )
        ]

        if remainder_expense <= 0:
            return split_rows

        if _is_fish_buchungstext(text_lower):
            split_rows.append(
                replace(
                    row,
                    umsatz_euro=-remainder_expense,
                    bu_gkto=BU_GKTO_FISCH_FOOD,
                    beleg_1="",
                    buchungstext=BOOKING_TEXT_FISHFOOD_EW_7,
                )
            )
            return split_rows

        if BOOKING_TEXT_BANK_DEPOSIT.lower() in text_lower or text_lower == BOOKING_TEXT_AN_BANK.lower():
            split_rows.append(
                replace(
                    row,
                    umsatz_euro=-remainder_expense,
                    bu_gkto=BU_GKTO_ALLOPAY_EXPENSE,
                    beleg_1="",
                    buchungstext=BOOKING_TEXT_AN_BANK,
                )
            )
            return split_rows

        return None

    def _assign_final_beleg(
        self,
        row: BuchungRow,
        safe_allopay_beleg_by_date: dict[str, str],
    ) -> BuchungRow:
        text_lower = _buchungstext_lower(row.buchungstext)
        if text_lower == "allo pay":
            return replace(
                row,
                beleg_1=safe_allopay_beleg_by_date.get(row.datum, ""),
                buchungstext=f"alloPay {row.datum}",
            )
        if "allo" in text_lower:
            return replace(row, beleg_1=safe_allopay_beleg_by_date.get(row.datum, ""))
        return replace(row, beleg_1="")

    def run(
        self,
        input_path: Path,
        output_path: Path,
        pdf_base_dir: Path | None = None,
        sheet_name: str | None = None,
    ) -> PipelineResult:
        if pdf_base_dir is None:
            pdf_base_dir = input_path.parent

        self.umsatz_rows = read_cashbook_rows(input_path, sheet_name=sheet_name)
        allopay_pdf_data = read_allopay_pdf_data(pdf_base_dir)
        self.allopay_rows = build_allopay_rows(allopay_pdf_data)
        self.final_rows = self.merge_final_rows(
            self.umsatz_rows,
            self.allopay_rows,
            {
                item.datum: item.allopay_payment_sum
                for item in allopay_pdf_data
                if item.datum != "" and item.allopay_payment_sum > 0
            },
        )

        saved_path, umsatz_count, allopay_count, final_count = build_workbook(
            self.umsatz_rows,
            self.allopay_rows,
            self.final_rows,
            output_path,
        )

        return PipelineResult(
            saved_path=saved_path,
            umsatz_count=umsatz_count,
            allopay_count=allopay_count,
            final_count=final_count,
            pdf_base_dir=pdf_base_dir,
        )
