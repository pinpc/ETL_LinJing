from __future__ import annotations

from dataclasses import dataclass, replace
from datetime import datetime
from decimal import Decimal, InvalidOperation
import os
from pathlib import Path
import re
import sys
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import pdfplumber

TARGET_COLUMNS = [
    "Umsatz Euro",
    "BU Gkto",
    "Beleg 1",
    "Datum",
    "KOST 1",
    "Bank",
    "Buchungstext",
]

STANDARD_KOST = "1000"
STANDARD_BANK = "1000"

BU_GKTO_POSITIVE = "8000"
BU_GKTO_NEGATIVE = "1360"
BU_GKTO_ALLOPAY_19 = "8400"
BU_GKTO_ALLOPAY_7 = "8300"
BU_GKTO_TIPS_0 = "4140"

BOOKING_TEXT_ALLO_PAY = "AllO Pay"
BOOKING_TEXT_TIPS = "Trinkgeld"
BOOKING_TEXT_TIPS_0 = "Umsatz 0 %"
BOOKING_TEXT_UMSATZ_19 = "Umsatz 19 %"
BOOKING_TEXT_UMSATZ_7 = "Umsatz 7 %"
BOOKING_TEXT_BANK = "an Bank"

MERGE_TOLERANCE = Decimal("0.01")
XL_EURO_NUM_FMT = "#,##0.00"


@dataclass(frozen=True)
class BuchungRow:
    umsatz_euro: Decimal
    bu_gkto: str
    beleg_1: str
    datum: str
    kost_1: str
    bank: str
    buchungstext: str


@dataclass(frozen=True)
class PipelineResult:
    saved_path: Path
    buchung_count: int
    allopay_count: int
    final_count: int
    pdf_base_dir: Path


@dataclass(frozen=True)
class AllopayPdfData:
    datum: str
    z_nummer: str
    umsatz_7: Decimal
    umsatz_19: Decimal
    dateiname: str


def as_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def norm_header(value: Any) -> str:
    return as_text(value)


def parse_amount(value: Any) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return Decimal("0")

    text = str(value).strip()
    if text.count(",") == 1 and text.count(".") >= 1:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")

    if text == "":
        return Decimal("0")

    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def convert_number(value: str) -> Decimal:
    text = (value or "").strip().replace("EUR", "").replace("€", "").strip()
    if text == "":
        return Decimal("0")

    if "," in text and "." in text:
        text = text.replace(",", "")
    elif "," in text:
        parts = text.split(",")
        if len(parts) == 2 and len(parts[1]) == 2:
            text = text.replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "." in text:
        parts = text.split(".")
        if len(parts) == 2 and len(parts[1]) != 2:
            text = text.replace(".", "")

    try:
        return Decimal(str(float(text)))
    except Exception:
        return Decimal("0")


def parse_date(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()

    text = str(value).strip()
    if text == "":
        return ""

    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass

    return text


def format_buchungstext_date_short(value: str) -> str:
    text = as_text(value)
    if text == "":
        return ""

    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y/%m/%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%d.%m.%y")
        except ValueError:
            pass

    return text


def format_beleg(value: Any) -> str:
    text = as_text(value)
    if text == "":
        return ""

    if text.upper().startswith("Z") and any(ch.isdigit() for ch in text):
        digits = "".join(ch for ch in text if ch.isdigit())
        if digits:
            return f"Z{int(digits):03d}"

    digits = "".join(ch for ch in text if ch.isdigit())
    if digits == "":
        return text.upper() if text.upper().startswith("Z") else text
    return f"Z{int(digits):03d}"


def date_sort_key(value: str) -> tuple[int, int, int, str]:
    text = as_text(value)
    if text == "":
        return (9999, 12, 31, text)

    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y/%m/%d", "%d/%m/%Y"):
        try:
            dt = datetime.strptime(text, fmt)
            return (dt.year, dt.month, dt.day, text)
        except ValueError:
            pass

    return (9999, 12, 31, text)


def sort_rows_by_date(rows: Iterable[BuchungRow]) -> list[BuchungRow]:
    return sorted(rows, key=lambda row: (date_sort_key(row.datum), row.beleg_1, row.buchungstext))


def trinkgeld_sum_by_date(buchung_rows: Iterable[BuchungRow]) -> dict[tuple[int, int, int], Decimal]:
    totals: dict[tuple[int, int, int], Decimal] = {}
    for row in buchung_rows:
        if row.buchungstext.strip().lower() != BOOKING_TEXT_TIPS.lower():
            continue
        key = date_sort_key(row.datum)[:3]
        totals[key] = totals.get(key, Decimal("0")) + row.umsatz_euro
    return totals


def is_allopay_income_row(row: BuchungRow) -> bool:
    return row.umsatz_euro > 0 and BOOKING_TEXT_ALLO_PAY.lower() in row.buchungstext.lower()


def build_trinkgeld_final_row(
    trinkgeld_net: Decimal,
    allopay_rows_for_date: list[BuchungRow],
    income_row: BuchungRow,
) -> BuchungRow | None:
    if trinkgeld_net >= Decimal("0"):
        return None

    template = allopay_rows_for_date[0] if allopay_rows_for_date else income_row
    return BuchungRow(
        umsatz_euro=(-trinkgeld_net).quantize(Decimal("0.01")),
        bu_gkto=BU_GKTO_TIPS_0,
        beleg_1=template.beleg_1 or income_row.beleg_1,
        datum=template.datum or income_row.datum,
        kost_1=STANDARD_KOST,
        bank=STANDARD_BANK,
        buchungstext=BOOKING_TEXT_TIPS_0,
    )


def _warn_allopay_merge_failure(message: str) -> None:
    print(f"WARNUNG [Blatt Final]: Allopay-Auflösung nicht mergebar {message}", file=sys.stderr)


def _find_header_row(ws, min_nonempty: int = 4, scan_rows: int = 30) -> int:
    best_row = 1
    best_score = -1

    for row_idx in range(1, min(ws.max_row, scan_rows) + 1):
        values = [norm_header(cell.value) for cell in ws[row_idx]]
        nonempty = sum(1 for value in values if value != "")
        score = nonempty

        wanted = {"datum", "date", "umsatz", "amount", "betrag", "金额", "交易日期", "识别号", "主题"}
        if any(value in wanted or value.lower() in wanted for value in values):
            score += 5
        if any(value == "金额" for value in values):
            score += 5
        if any(value == "交易日期" for value in values):
            score += 5

        raw_values = [cell.value for cell in ws[row_idx]]
        numericish = sum(
            1
            for value in raw_values
            if isinstance(value, (int, float, Decimal, datetime))
            or (isinstance(value, str) and value.strip().lstrip("-").replace(".", "").replace(",", "").isdigit())
        )
        if numericish >= 2:
            score -= 4

        if score > best_score and nonempty >= min_nonempty:
            best_score = score
            best_row = row_idx

    return best_row


def _map_columns(headers: list[str]) -> dict[str, int]:
    exact = {header: idx for idx, header in enumerate(headers)}
    lower = {header.lower(): idx for idx, header in enumerate(headers)}

    def pick(*candidates: str) -> int | None:
        for candidate in candidates:
            if candidate in exact:
                return exact[candidate]
        for candidate in candidates:
            if candidate.lower() in lower:
                return lower[candidate.lower()]
        for candidate in candidates:
            token = candidate.lower()
            for header, idx in lower.items():
                if token in header:
                    return idx
        return None

    mapping: dict[str, int] = {}
    mapping["Umsatz Euro"] = pick("金额", "Umsatz Euro", "Umsatz", "Amount", "Betrag")  # type: ignore[assignment]
    mapping["Beleg 1"] = pick("识别号", "Beleg 1", "Beleg", "Receipt", "Bon")  # type: ignore[assignment]
    mapping["Datum"] = pick("交易日期", "Datum", "Date")  # type: ignore[assignment]
    mapping["主题"] = pick("主题", "Subject", "Betreff")  # type: ignore[assignment]
    return {key: value for key, value in mapping.items() if value is not None}  # type: ignore[misc]


def read_cashbook_rows(input_xlsx: Path, sheet_name: str = "cashbook") -> list[BuchungRow]:
    workbook = load_workbook(input_xlsx, data_only=True)

    wanted = sheet_name.strip()
    name_map = {name.lower(): name for name in workbook.sheetnames}
    resolved = name_map.get(wanted.lower(), wanted)
    if resolved not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}")

    worksheet = workbook[resolved]
    header_row = _find_header_row(worksheet)
    headers = [norm_header(cell.value) for cell in worksheet[header_row]]
    col_map = _map_columns(headers)

    rows: list[BuchungRow] = []
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        values = [cell.value for cell in worksheet[row_idx]]
        if all(value is None or str(value).strip() == "" for value in values):
            continue

        def get(column_name: str) -> Any:
            idx = col_map.get(column_name)
            return values[idx] if idx is not None and idx < len(values) else None

        amount = parse_amount(get("Umsatz Euro"))
        datum = parse_date(get("Datum"))
        if datum == "" and amount == 0:
            continue

        beleg = format_beleg(get("Beleg 1"))
        subject = as_text(get("主题"))
        if beleg.upper().startswith("Z"):
            date_suffix = format_buchungstext_date_short(datum)
            buchungstext = (
                f"{BOOKING_TEXT_ALLO_PAY} {date_suffix}" if date_suffix else BOOKING_TEXT_ALLO_PAY
            )
        else:
            buchungstext = subject
        bu_gkto = BU_GKTO_POSITIVE if amount >= 0 else BU_GKTO_NEGATIVE

        rows.append(
            BuchungRow(
                umsatz_euro=amount,
                bu_gkto=bu_gkto,
                beleg_1=beleg,
                datum=datum,
                kost_1=STANDARD_KOST,
                bank=STANDARD_BANK,
                buchungstext=buchungstext,
            )
        )

    return rows


def _find_allopay_pdfs(base_path: Path) -> list[Path]:
    pattern = re.compile(r"asia[ _]\d{2}[-_.]\d{2}[-_.]\d{4}\.pdf$", re.IGNORECASE)
    matched: list[Path] = []

    for root, _dirs, files in os.walk(base_path):
        for filename in files:
            if not filename.lower().endswith(".pdf"):
                continue
            if pattern.match(filename):
                matched.append(Path(root) / filename)

    return sorted(matched)


def _extract_allopay_from_pdf(pdf_path: Path) -> AllopayPdfData:
    umsatz_7 = Decimal("0")
    umsatz_19 = Decimal("0")
    datum = ""
    z_nummer = ""

    with pdfplumber.open(str(pdf_path)) as pdf:
        full_text = ""
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                full_text += text + "\n"

        z_match = re.search(r"Z-Nummer\s*\(Berichtnummer\):\s*(\d+)", full_text, re.IGNORECASE)
        if z_match:
            z_nummer = z_match.group(1)

        gen_match = re.search(r"Generiert um:\s*(\d{2})-(\d{2})-(\d{4})", full_text)
        if gen_match:
            datum = f"{gen_match.group(1)}.{gen_match.group(2)}.{gen_match.group(3)}"
        else:
            name_match = re.search(r"asia[ _](\d{2})-(\d{2})-(\d{4})\.pdf", pdf_path.name, re.IGNORECASE)
            if name_match:
                datum = f"{name_match.group(1)}.{name_match.group(2)}.{name_match.group(3)}"

        for page in pdf.pages:
            for table in page.extract_tables() or []:
                for table_row in table or []:
                    if not table_row:
                        continue
                    row_text = " ".join(str(cell) if cell else "" for cell in table_row)
                    if "Umsatz 7%" in row_text or "Umsatz 7 %" in row_text:
                        numbers = re.findall(r"(\d+(?:[.,]\d+)?)", row_text)
                        if numbers:
                            umsatz_7 = convert_number(numbers[-1])
                    if "Umsatz 19%" in row_text or "Umsatz 19 %" in row_text:
                        numbers = re.findall(r"(\d+(?:[.,]\d+)?)", row_text)
                        if numbers:
                            umsatz_19 = convert_number(numbers[-1])

        if umsatz_7 == 0:
            match_7 = re.search(
                r"Umsatz\s*7\s*%\s*(?:€\s*)?([\d.,]+)\s*(?:€\s*)?([\d.,]+)\s*(?:€\s*)?([\d.,]+)",
                full_text,
            )
            if match_7:
                umsatz_7 = convert_number(match_7.group(3))

        if umsatz_19 == 0:
            match_19 = re.search(
                r"Umsatz\s*19\s*%\s*(?:€\s*)?([\d.,]+)\s*(?:€\s*)?([\d.,]+)\s*(?:€\s*)?([\d.,]+)",
                full_text,
            )
            if match_19:
                umsatz_19 = convert_number(match_19.group(3))

    return AllopayPdfData(
        datum=datum,
        z_nummer=z_nummer,
        umsatz_7=umsatz_7,
        umsatz_19=umsatz_19,
        dateiname=pdf_path.name,
    )


def read_allopay_rows(pdf_base_dir: Path) -> list[BuchungRow]:
    rows: list[BuchungRow] = []

    for pdf_path in _find_allopay_pdfs(pdf_base_dir):
        data = _extract_allopay_from_pdf(pdf_path)
        if as_text(data.datum) == "":
            continue

        beleg = f"Z{str(data.z_nummer).zfill(3)}" if data.z_nummer else "Z000"

        if data.umsatz_19 > 0:
            rows.append(
                BuchungRow(
                    umsatz_euro=data.umsatz_19,
                    bu_gkto=BU_GKTO_ALLOPAY_19,
                    beleg_1=beleg,
                    datum=data.datum,
                    kost_1=STANDARD_KOST,
                    bank=STANDARD_BANK,
                    buchungstext=BOOKING_TEXT_UMSATZ_19,
                )
            )

        if data.umsatz_7 > 0:
            rows.append(
                BuchungRow(
                    umsatz_euro=data.umsatz_7,
                    bu_gkto=BU_GKTO_ALLOPAY_7,
                    beleg_1=beleg,
                    datum=data.datum,
                    kost_1=STANDARD_KOST,
                    bank=STANDARD_BANK,
                    buchungstext=BOOKING_TEXT_UMSATZ_7,
                )
            )

    return sort_rows_by_date(rows)


def _load_or_create_workbook(output_xlsx: Path):
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    if output_xlsx.exists():
        return load_workbook(output_xlsx)

    workbook = Workbook()
    if workbook.sheetnames == ["Sheet"]:
        workbook.remove(workbook["Sheet"])
    return workbook


def _write_rows_sheet(workbook, rows: Iterable[BuchungRow], sheet_name: str) -> int:
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])

    worksheet = workbook.create_sheet(sheet_name)
    worksheet.append(TARGET_COLUMNS)

    row_count = 0
    for row in rows:
        worksheet.append(
            [
                float(row.umsatz_euro),
                row.bu_gkto,
                row.beleg_1,
                row.datum,
                row.kost_1,
                row.bank,
                row.buchungstext,
            ]
        )
        row_count += 1

    if worksheet.max_row >= 2:
        sum_row_idx = worksheet.max_row + 1
        worksheet.cell(row=sum_row_idx, column=1).value = f"=SUM(A2:A{worksheet.max_row})"
        worksheet.cell(row=sum_row_idx, column=1).number_format = XL_EURO_NUM_FMT
        worksheet.cell(row=sum_row_idx, column=7).value = "Summe"
        for col_idx in range(1, len(TARGET_COLUMNS) + 1):
            worksheet.cell(row=sum_row_idx, column=col_idx).font = Font(bold=True)

    for row_idx in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row_idx, column=1).number_format = XL_EURO_NUM_FMT

    for col_idx, col_name in enumerate(TARGET_COLUMNS, start=1):
        max_len = len(col_name)
        for values in worksheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
            value = values[0]
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 60)

    return row_count


def _rename_umsatz_bank_header(workbook) -> None:
    if "Umsatz" not in workbook.sheetnames:
        return
    worksheet = workbook["Umsatz"]
    for cell in worksheet[1]:
        if norm_header(cell.value) == "Kasse":
            cell.value = "Bank"


def _save_workbook(workbook, output_xlsx: Path) -> Path:
    try:
        workbook.save(output_xlsx)
        return output_xlsx
    except PermissionError:
        candidates = [output_xlsx.with_name(output_xlsx.stem + "_new" + output_xlsx.suffix)]
        candidates.extend(
            output_xlsx.with_name(output_xlsx.stem + f"_new_{idx}" + output_xlsx.suffix)
            for idx in range(2, 11)
        )

        for alt_path in candidates:
            try:
                workbook.save(alt_path)
                print(f"WARNING: Could not overwrite (file open?). Saved to: {alt_path}")
                return alt_path
            except PermissionError:
                continue

        raise


def build_workbook(
    buchung_rows: list[BuchungRow],
    allopay_rows: list[BuchungRow],
    final_rows: list[BuchungRow],
    output_xlsx: Path,
) -> tuple[Path, int, int, int]:
    workbook = _load_or_create_workbook(output_xlsx)
    buchung_count = _write_rows_sheet(workbook, buchung_rows, "Umsatz")
    allopay_count = _write_rows_sheet(workbook, allopay_rows, "Allopay")
    final_count = _write_rows_sheet(workbook, final_rows, "Final")
    _rename_umsatz_bank_header(workbook)
    saved_path = _save_workbook(workbook, output_xlsx)
    return saved_path, buchung_count, allopay_count, final_count


class AsiaKasseETL:
    def __init__(self) -> None:
        self.buchung_rows: list[BuchungRow] = []
        self.allopay_rows: list[BuchungRow] = []
        self.final_rows: list[BuchungRow] = []

    def merge_final_rows(self, buchung_rows: list[BuchungRow], allopay_rows: list[BuchungRow]) -> list[BuchungRow]:
        def normalize_final_text(row: BuchungRow) -> BuchungRow:
            if row.buchungstext.strip().lower() == "bankeinzahlung":
                return replace(row, buchungstext=BOOKING_TEXT_BANK)
            return row

        if not buchung_rows:
            return []

        if not allopay_rows:
            for row in buchung_rows:
                if is_allopay_income_row(row):
                    _warn_allopay_merge_failure(
                        f"({row.datum}): keine Allopay-PDF-Zeilen geladen; "
                        f"Kasse bleibt {row.umsatz_euro} EUR."
                    )
            return sort_rows_by_date([normalize_final_text(row) for row in buchung_rows])

        allopay_by_date: dict[tuple[int, int, int], list[BuchungRow]] = {}
        allopay_sum_by_date: dict[tuple[int, int, int], Decimal] = {}
        trinkgeld_by_date = trinkgeld_sum_by_date(buchung_rows)

        for row in allopay_rows:
            key = date_sort_key(row.datum)[:3]
            allopay_by_date.setdefault(key, []).append(row)
            allopay_sum_by_date[key] = allopay_sum_by_date.get(key, Decimal("0")) + row.umsatz_euro

        final_rows: list[BuchungRow] = []
        for row in buchung_rows:
            key = date_sort_key(row.datum)[:3]

            if is_allopay_income_row(row):
                pdf_umsatz_sum = allopay_sum_by_date.get(key)
                if pdf_umsatz_sum is not None:
                    trinkgeld = trinkgeld_by_date.get(key, Decimal("0"))
                    adjusted_income = (row.umsatz_euro + trinkgeld).quantize(Decimal("0.01"))
                    diff = abs(pdf_umsatz_sum - adjusted_income)
                    if diff <= MERGE_TOLERANCE:
                        final_rows.extend(
                            normalize_final_text(allopay_row) for allopay_row in allopay_by_date[key]
                        )
                        trinkgeld_row = build_trinkgeld_final_row(trinkgeld, allopay_by_date[key], row)
                        if trinkgeld_row is not None:
                            final_rows.append(trinkgeld_row)
                        continue

                    tip_hint = (
                        f", Trinkgeld {trinkgeld} EUR, netto {adjusted_income} EUR"
                        if trinkgeld != Decimal("0")
                        else ""
                    )
                    _warn_allopay_merge_failure(
                        f"({row.datum}): Kasse {row.umsatz_euro} EUR{tip_hint} vs. Summe PDF "
                        f"({BOOKING_TEXT_UMSATZ_19} + {BOOKING_TEXT_UMSATZ_7}) {pdf_umsatz_sum} EUR "
                        f"(|Diff|={diff}, Toleranz {MERGE_TOLERANCE})."
                    )
                else:
                    _warn_allopay_merge_failure(
                        f"({row.datum}): keine PDF-Umsatzzeilen (19 %/7 %) für dieses Datum; "
                        f"Kasse bleibt {row.umsatz_euro} EUR."
                    )

            final_rows.append(normalize_final_text(row))

        return sort_rows_by_date(final_rows)

    def run(
        self,
        input_path: Path,
        output_path: Path,
        pdf_base_dir: Path | None = None,
        sheet_name: str = "cashbook",
    ) -> PipelineResult:
        if pdf_base_dir is None:
            pdf_base_dir = input_path.parents[2]

        self.buchung_rows = read_cashbook_rows(input_path, sheet_name=sheet_name)
        self.allopay_rows = read_allopay_rows(pdf_base_dir)
        self.final_rows = self.merge_final_rows(self.buchung_rows, self.allopay_rows)

        saved_path, buchung_count, allopay_count, final_count = build_workbook(
            self.buchung_rows,
            self.allopay_rows,
            self.final_rows,
            output_path,
        )

        return PipelineResult(
            saved_path=saved_path,
            buchung_count=buchung_count,
            allopay_count=allopay_count,
            final_count=final_count,
            pdf_base_dir=pdf_base_dir,
        )


def run_cli(
    input_path: Path,
    output_path: Path,
    pdf_base_dir: Path | None = None,
    sheet_name: str = "cashbook",
) -> PipelineResult:
    return AsiaKasseETL().run(input_path, output_path, pdf_base_dir=pdf_base_dir, sheet_name=sheet_name)
