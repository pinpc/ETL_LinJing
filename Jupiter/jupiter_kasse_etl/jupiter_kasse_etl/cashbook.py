from __future__ import annotations

import re
from decimal import Decimal
from pathlib import Path
from typing import Any

import pdfplumber
from openpyxl import load_workbook

from .config import (
    BU_GKTO_ALLOPAY_EXPENSE,
    BU_GKTO_BAUMARKT_EXPENSE,
    BU_GKTO_INCOME,
    STANDARD_BANK,
    STANDARD_KOST,
)
from .models import BuchungRow, CashbookTransaction
from .utils import convert_german_number, parse_date, sort_rows_by_date


def _contains_any(text: str, tokens: tuple[str, ...]) -> bool:
    lowered = text.lower()
    return any(token in lowered for token in tokens)


def _find_header_row(ws, scan_rows: int = 30) -> int:
    best_row = 1
    best_score = -1

    for row_idx in range(1, min(ws.max_row, scan_rows) + 1):
        values = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[row_idx]]
        row_text = " ".join(values).lower()
        nonempty = sum(1 for value in values if value)
        score = nonempty

        if _contains_any(row_text, ("datum", "beleg datum", "einnahmen", "ausgaben", "geschäft", "收入", "支出")):
            score += 8

        if score > best_score and nonempty >= 3:
            best_score = score
            best_row = row_idx

    return best_row


def _map_excel_columns(headers: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}

    for idx, header in enumerate(headers):
        value = header.lower()
        if "datum" in value or "日期" in value:
            mapping.setdefault("datum", idx)
        elif "einnahmen" in value or "rechts" in value or "收入" in value:
            mapping.setdefault("einnahmen", idx)
        elif "ausgaben" in value or "links" in value or "支出" in value:
            mapping.setdefault("ausgaben", idx)
        elif "geschäft" in value or "用途" in value:
            mapping.setdefault("text", idx)

    return mapping


def _should_stop_row(values: list[Any]) -> bool:
    row_text = " ".join(str(value) for value in values if value is not None).lower()
    return _contains_any(row_text, ("summe", "unterschrift", "bestand", "收入总额"))


def _read_excel_transactions(input_path: Path, sheet_name: str | None = None) -> list[CashbookTransaction]:
    workbook = load_workbook(input_path, data_only=True)

    if sheet_name:
        name_map = {name.lower(): name for name in workbook.sheetnames}
        resolved = name_map.get(sheet_name.lower(), sheet_name)
        if resolved not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}")
        worksheet = workbook[resolved]
    else:
        worksheet = next(
            workbook[name] for name in workbook.sheetnames if not name.startswith("__")
        )

    header_row = _find_header_row(worksheet)
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in worksheet[header_row]]
    col_map = _map_excel_columns(headers)

    transactions: list[CashbookTransaction] = []
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        values = [cell.value for cell in worksheet[row_idx]]
        if all(value is None or str(value).strip() == "" for value in values):
            continue
        if _should_stop_row(values):
            break

        def get(column_name: str) -> Any:
            idx = col_map.get(column_name)
            return values[idx] if idx is not None and idx < len(values) else None

        datum = parse_date(get("datum"))
        einnahmen = convert_german_number(get("einnahmen"))
        ausgaben = convert_german_number(get("ausgaben"))
        text = str(get("text")).strip() if get("text") is not None else ""

        if datum == "":
            continue
        if einnahmen <= 0 and ausgaben <= 0:
            continue

        transactions.append(
            CashbookTransaction(
                datum=datum,
                einnahmen=einnahmen,
                ausgaben=ausgaben,
                buchungstext=text,
            )
        )

    return transactions


def _read_pdf_transactions(pdf_path: Path) -> list[CashbookTransaction]:
    transactions: list[CashbookTransaction] = []

    with pdfplumber.open(str(pdf_path)) as pdf:
        page = pdf.pages[0]
        tables = page.extract_tables()
        if not tables:
            return []

        table = tables[0]
        header_row_idx = -1
        col_datum_idx = -1
        col_einnahmen_idx = -1
        col_ausgaben_idx = -1
        col_text_idx = -1

        for row_idx, row in enumerate(table):
            row_text = str(row).lower()
            if "beleg" in row_text or "datum" in row_text or "einnahmen" in row_text:
                header_row_idx = row_idx
                for col_idx, cell in enumerate(row):
                    if cell is None:
                        continue
                    cell_text = str(cell).lower()
                    if "datum" in cell_text:
                        col_datum_idx = col_idx
                    elif "einnahmen" in cell_text or "rechts" in cell_text or "收入" in cell_text:
                        col_einnahmen_idx = col_idx
                    elif "ausgaben" in cell_text or "links" in cell_text or "支出" in cell_text:
                        col_ausgaben_idx = col_idx
                    elif "geschäfts" in cell_text or "用途" in cell_text:
                        col_text_idx = col_idx
                break

        if header_row_idx == -1:
            return []

        if col_datum_idx == -1:
            col_datum_idx = 1
        if col_einnahmen_idx == -1:
            col_einnahmen_idx = 2
        if col_ausgaben_idx == -1:
            col_ausgaben_idx = 3
        if col_text_idx == -1:
            col_text_idx = 5

        for row in table[header_row_idx + 1 :]:
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            row_text = str(row).lower()
            if _contains_any(row_text, ("summe", "unterschrift", "bestand", "收入总额")):
                break

            datum = ""
            if col_datum_idx < len(row) and row[col_datum_idx]:
                datum = parse_date(str(row[col_datum_idx]).strip())

            if not re.match(r"\d{2}\.\d{2}\.\d{4}", datum):
                continue

            einnahmen = (
                convert_german_number(row[col_einnahmen_idx])
                if col_einnahmen_idx < len(row)
                else Decimal("0")
            )
            ausgaben = (
                convert_german_number(row[col_ausgaben_idx])
                if col_ausgaben_idx < len(row)
                else Decimal("0")
            )
            text = str(row[col_text_idx]).strip() if col_text_idx < len(row) and row[col_text_idx] else ""

            if einnahmen > 0 or ausgaben > 0:
                transactions.append(
                    CashbookTransaction(
                        datum=datum,
                        einnahmen=einnahmen,
                        ausgaben=ausgaben,
                        buchungstext=text,
                    )
                )

    return transactions


def get_bu_gkto(umsatz: Decimal, text: str) -> int:
    text_lower = text.lower()

    if umsatz < 0:
        if "allo" in text_lower or "bankeinzahlung" in text_lower:
            return BU_GKTO_ALLOPAY_EXPENSE
        if "baumarkt" in text_lower or "v-baumarkt" in text_lower:
            return BU_GKTO_BAUMARKT_EXPENSE
        return BU_GKTO_BAUMARKT_EXPENSE
    return BU_GKTO_INCOME


def _build_buchung_rows(transactions: list[CashbookTransaction]) -> list[BuchungRow]:
    rows: list[BuchungRow] = []

    for transaction in transactions:
        if transaction.einnahmen > 0:
            rows.append(
                BuchungRow(
                    umsatz_euro=transaction.einnahmen.quantize(Decimal("0.01")),
                    bu_gkto=get_bu_gkto(transaction.einnahmen, transaction.buchungstext),
                    beleg_1="",
                    datum=transaction.datum,
                    kost_1=STANDARD_KOST,
                    bank=STANDARD_BANK,
                    buchungstext=transaction.buchungstext,
                )
            )
        if transaction.ausgaben > 0:
            negative_amount = (-transaction.ausgaben).quantize(Decimal("0.01"))
            rows.append(
                BuchungRow(
                    umsatz_euro=negative_amount,
                    bu_gkto=get_bu_gkto(negative_amount, transaction.buchungstext),
                    beleg_1="",
                    datum=transaction.datum,
                    kost_1=STANDARD_KOST,
                    bank=STANDARD_BANK,
                    buchungstext=transaction.buchungstext,
                )
            )

    sorted_rows = sort_rows_by_date(rows)
    return [
        BuchungRow(
            umsatz_euro=row.umsatz_euro,
            bu_gkto=row.bu_gkto,
            beleg_1=f"Z{idx:03d}",
            datum=row.datum,
            kost_1=row.kost_1,
            bank=row.bank,
            buchungstext=row.buchungstext,
        )
        for idx, row in enumerate(sorted_rows, start=1)
    ]


def read_cashbook_rows(input_path: Path, sheet_name: str | None = None) -> list[BuchungRow]:
    suffix = input_path.suffix.lower()
    if suffix == ".pdf":
        transactions = _read_pdf_transactions(input_path)
    elif suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        transactions = _read_excel_transactions(input_path, sheet_name=sheet_name)
    else:
        raise ValueError(f"Unsupported cashbook format: {input_path.suffix}")

    return _build_buchung_rows(transactions)
