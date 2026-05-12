from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .config import COLUMN_WIDTHS, TARGET_COLUMNS, XL_EURO_NUM_FMT
from .models import BuchungRow


def _append_sheet(workbook: Workbook, rows: list[BuchungRow], sheet_name: str) -> int:
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.append(TARGET_COLUMNS)

    for row in rows:
        worksheet.append(
            [
                float(row.umsatz_euro),
                row.bu_gkto,
                row.beleg_1,
                row.datum,
                row.kost_1,
                row.bank,
                row.buchungstext,
            ]
        )

    if rows:
        worksheet.append(
            [
                float(sum((row.umsatz_euro for row in rows), Decimal("0"))),
                "",
                "",
                "",
                "",
                "",
                "Gesamtbetrag",
            ]
        )

    _apply_excel_formatting(worksheet)
    return len(rows)


def _apply_excel_formatting(worksheet) -> None:
    max_row = worksheet.max_row
    max_col = len(TARGET_COLUMNS)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx in range(2, max_row + 1):
        amount_cell = worksheet.cell(row=row_idx, column=1)
        if amount_cell.value not in (None, ""):
            amount_cell.number_format = XL_EURO_NUM_FMT
            try:
                amount_value = float(amount_cell.value)
            except Exception:
                amount_value = None
            if amount_value is not None and amount_value < 0:
                amount_cell.font = Font(color="FF0000")

        if worksheet.cell(row=row_idx, column=7).value == "Gesamtbetrag":
            for col_idx in range(1, max_col + 1):
                worksheet.cell(row=row_idx, column=col_idx).font = Font(bold=True)

    for col_letter, width in COLUMN_WIDTHS.items():
        worksheet.column_dimensions[col_letter].width = width


def _save_workbook(workbook: Workbook, output_xlsx: Path) -> Path:
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)

    try:
        workbook.save(output_xlsx)
        return output_xlsx
    except PermissionError:
        candidates = [output_xlsx.with_name(output_xlsx.stem + "_new" + output_xlsx.suffix)]
        candidates.extend(
            output_xlsx.with_name(output_xlsx.stem + f"_new_{idx}" + output_xlsx.suffix)
            for idx in range(2, 11)
        )

        for alt_path in candidates:
            try:
                workbook.save(alt_path)
                print(f"WARNING: Could not overwrite (file open?). Saved to: {alt_path}")
                return alt_path
            except PermissionError:
                continue

        raise


def build_workbook(
    umsatz_rows: list[BuchungRow],
    allopay_rows: list[BuchungRow],
    final_rows: list[BuchungRow],
    output_xlsx: Path,
) -> tuple[Path, int, int, int]:
    workbook = Workbook()
    if workbook.sheetnames == ["Sheet"]:
        workbook.remove(workbook["Sheet"])

    umsatz_count = _append_sheet(workbook, umsatz_rows, "Umsatz")
    allopay_count = _append_sheet(workbook, allopay_rows, "Allopay")
    final_count = _append_sheet(workbook, final_rows, "Final")

    saved_path = _save_workbook(workbook, output_xlsx)
    return saved_path, umsatz_count, allopay_count, final_count
